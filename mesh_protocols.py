import openpyxl
from pathlib import Path
import xml.etree.ElementTree as ET


# 標高モデルxmlの解析 参考-> https://pycra.hatenablog.com/entry/2017/09/17/221026

XML_FILEPATH1 = Path.cwd() / r'xmls/FG-GML-5437-27-77-DEM5A-20161001.xml'
XML_FILEPATH2 = Path.cwd() / r'xmls/FG-GML-5437-27-87-DEM5A-20161001.xml'
XML_FILEPATH3 = Path.cwd() / r'xmls/FG-GML-5437-27-78-DEM5A-20161001.xml'
XML_FILEPATH4 = Path.cwd() / r'xmls/FG-GML-5437-27-88-DEM5A-20161001.xml'
DEM_EXCELPATH = Path.cwd() / r'mesh_dem/mesh_dem_data.xlsx'

XML_FILEPATH_LIST = [XML_FILEPATH1, XML_FILEPATH2, XML_FILEPATH3, XML_FILEPATH4]

def extract_DEM_from_xml(xml_filename):
    xml = xml_filename
    tree = ET.parse(xml)
    root = tree.getroot()
    target = root.find('.//{http://www.opengis.net/gml/3.2}tupleList') # このURL部分は'gml'と記述されている部分。# 参考-> https://nixeneko.hatenablog.com/entry/2022/04/02/233241
    grid_envelope_high:list = root.find('.//{http://www.opengis.net/gml/3.2}GridEnvelope/{http://www.opengis.net/gml/3.2}high')\
                        .text.split(' ') # 5mメッシュの横縦に並んでいる数を取得。[横の数-1, 縦の数-1]になっている。
    meshid_gsi = root.find('.//{http://fgd.gsi.go.jp/spec/2008/FGD_GMLSchema}DEM/{http://fgd.gsi.go.jp/spec/2008/FGD_GMLSchema}mesh')\
                        .text # 第三次メッシュのidを取得
    # start_and_stop = root.find() 
    lowercorner_location = root.find('.//{http://www.opengis.net/gml/3.2}lowerCorner').text.split()
    lowercorner_location = tuple(map(float, lowercorner_location))
    uppercorner_location = root.find('.//{http://www.opengis.net/gml/3.2}upperCorner').text.split()
    uppercorner_location = tuple(map(float, uppercorner_location))
    # print(lowercorner_location)        
                        
    # print(grid_envelope_high)
    
    if target is not None:
        # targetがマッチした場合。
        lines = target.text.split() # textメソッドで、'tupleList'の中身を取得
        
        elevation_lis:list[float] = []
        for line in lines:
            value = line.split(',')[1]
            elevation_lis.append(float(value))
        
        elevation_lis = convert_1d_to_2d(elevation_lis, int(grid_envelope_high[0])+1)
        # print(elevation_lis)
        
        try:
            wb = openpyxl.load_workbook(DEM_EXCELPATH)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        try:
            ws = wb[str(meshid_gsi)]
        except KeyError:
            ws = wb.create_sheet(str(meshid_gsi))
        try:
            ws_base = wb[f'baseinfo_{str(meshid_gsi)}']
        except KeyError:
            ws_base = wb.create_sheet(f'baseinfo_{str(meshid_gsi)}')
        
        write_list_2d(ws, elevation_lis, 1, 1) # 標高データ書き込み
        ws_base['A1'] = uppercorner_location[0] # uppercornerの緯度
        ws_base['B1'] = uppercorner_location[1] # uppercornerの経度
        ws_base['A2'] = lowercorner_location[0]
        ws_base['B2'] = lowercorner_location[1]
        
        wb.save(DEM_EXCELPATH)
        
        # Mesh.third_mesh_infodict[int(meshid_gsi)] = [lowercorner_location, uppercorner_location]
        
    else:
        print("targetがマッチしていません。")
        
def start_extracting(filepathlis):
    for file in filepathlis:
        extract_DEM_from_xml(file)

def load_dem_file():
    return openpyxl.load_workbook(DEM_EXCELPATH)
    
        

def convert_1d_to_2d(l, cols): # 一次元リストを二次元配列に変換して返す関数。https://note.nkmk.me/python-list-ndarray-1d-to-2d/
    return [l[i:i + cols] for i in range(0, len(l), cols)]

def write_list_2d(sheet, l_2d, start_row, start_col):# 2次元配列をエクセルに書き込む
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

if __name__ == '__main__':
    start_extracting(XML_FILEPATH_LIST)
    
    # print(Mesh.third_mesh_infodict)

# print(lines)


