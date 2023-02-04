from pathlib import Path
import pandas as pd
import numpy as np
from decimal import Decimal
import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib


import datetime

# rainfall.txt やその他パラメータを設定するファイルをフォーマットする。


# rainfall


def format_rainfall(year, month, day):

    RAINFALL_XLSX = Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day:02}.xlsx'
    RAINFALL_FORMAT = Path.cwd().parent / f'rainfalls/rainfall_{year}{month:02}{day:02}.txt'

    f_txt = open(RAINFALL_FORMAT, 'w', newline='')
    
    wb:openpyxl.Workbook = openpyxl.load_workbook(RAINFALL_XLSX)
    ws = wb.worksheets[0]
    
    # 1行目
    a1 = ws['A1'].value
    
    first_string = f'{a1:>5}\n'
    f_txt.write(str(first_string))
    second_row_lis = [f'{1:>5}', ' '*58, str(626.697),' '*2 ,str(1149.770), '\n']
    second_strings = ''.join(second_row_lis)
    f_txt.write(str(second_strings))
    
    # 3行目以降
    all_lis = []
    for row in ws.iter_rows(min_row=3):
        temp_lis = []
        for cell in row:
            st = Decimal(cell.value)
            temp_lis.append(str(st))
        
        all_lis.append(temp_lis)
    
    for line in all_lis:
        n_row = list(map(Decimal, line))
        temp_lis2 = []
        for i in range(0, 7):
            temp_lis2.append(f'{n_row[i]:>10.2f}')
        temp_lis2.append(f'{n_row[7]:10.2f}\n')
        n_string = ''.join(temp_lis2)
        
        f_txt.write(str(n_string))
        
    f_txt.close()
    
def format_rainfall_extra(year, month, day1, day2):

    RAINFALL_XLSX = Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day1:02}-{day2:02}.xlsx'
    RAINFALL_FORMAT = Path.cwd().parent / f'rainfalls/rainfall_{year}{month:02}{day1:02}-{day2:02}.txt'

    f_txt = open(RAINFALL_FORMAT, 'w', newline='')
    
    wb:openpyxl.Workbook = openpyxl.load_workbook(RAINFALL_XLSX)
    ws = wb.worksheets[0]
    
    # 1行目
    a1 = ws['A1'].value
    
    first_string = f'{a1:>5}\n'
    f_txt.write(str(first_string))
    second_row_lis = [f'{1:>5}', ' '*58, str(626.697),' '*2 ,str(1149.770), '\n']
    second_strings = ''.join(second_row_lis)
    f_txt.write(str(second_strings))
    
    # 3行目以降
    all_lis = []
    for row in ws.iter_rows(min_row=3):
        temp_lis = []
        for cell in row:
            st = Decimal(cell.value)
            temp_lis.append(str(st))
        
        all_lis.append(temp_lis)
    
    for line in all_lis:
        n_row = list(map(Decimal, line))
        temp_lis2 = []
        for i in range(0, 7):
            temp_lis2.append(f'{n_row[i]:>10.2f}')
        temp_lis2.append(f'{n_row[7]:10.2f}\n')
        n_string = ''.join(temp_lis2)
        
        f_txt.write(str(n_string))
        
    f_txt.close()
    
def format_timedef(year, month, day, start_end_tuple:tuple, make_df=True, make_graph=True):
    # 時間定義ファイルの作成
    TIMEDEF_TXT = Path.cwd().parent / f'timedef/timedef_{year}{month:02}{day:02}.txt'
    f_timedef = open(TIMEDEF_TXT, 'w', newline='')
    
    # 開始時刻と終了時刻のデータ番号を計算する。
    start_hour = start_end_tuple[0]
    end_hour = start_end_tuple[1]
    if 1 <= (start_hour * 60 / 10) <= 143:
        start_data_num = int(start_hour * 60 / 10) # 00:10から数えた順番
    elif start_hour==0:
        start_data_num = 1
    else:
        print(f'start_data_num:{(start_hour+1) * 60 / 10}は0から143の間でなければならない')
        
    if 2<= (end_hour * 60 / 10) <= 144:
        end_data_num = int(end_hour * 60 / 10)
    else:
        (f'end_data_num:{end_hour * 60 / 10}は1から144の間でなければならない')
        
    
    
    
    first_row = f'{1:>5}\n'
    f_timedef.write(str(first_row))
    second_row = [f'{int(start_data_num):>5}', f'{int(end_data_num):>5}', f"{1:>10.3f}", f'{10:>5}', f"{year:>5}", f'{month:>5}', f'{day:>5}', f'{0:>5}',f'{9:>5}', f'{0:>5}\n']
    second_row = ''.join(second_row)
    f_timedef.write(str(second_row))
    third_row = [f'{10:>5}', f'{10:>5}', f'{1:>10}', f'{1:>5}']
    third_row = ''.join(third_row)
    f_timedef.write(str(third_row))
    
    f_timedef.close()
    
    
    # 降雨データのグラフ生成およびデータフレームの保存
    if make_df == True:
        RAINFALL_XLSX = Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day:02}.xlsx'
        rain_lis = []
        time_lis = []
        
        wb:openpyxl.Workbook = openpyxl.load_workbook(RAINFALL_XLSX)
        ws = wb.worksheets[0]
        
        count = 0
        for num in range(start_data_num, end_data_num+1):
            if (col_num := num%8) ==0:
                col_num = 8
            # else:
            #     col_num=col_num
            
            rain_lis.append(ws.cell(row=(((num - col_num)/8 + 1) +2), column=col_num).value)
            time_lis.append(count*10)
            count += 1
        rain_array = np.array(rain_lis)
        minute_array = np.array(time_lis)
        
        array_2d = np.stack([minute_array, rain_array], 1)
        # print(array_2d)
        df_rain = pd.DataFrame(array_2d, columns=['time', 'intensity'])
        # print(df_rain)
        RAIN_DF_PATH = Path.cwd() / f'precipitations/usage/rainfall(intensity)_usage_{year}{month:02}{day:02}.csv'
        
        df_rain.to_csv(RAIN_DF_PATH, header=True, index=False)
        
        # print(rain_array)
        # print(minute_array)
        if make_graph == True:
            # plt.bar(x=df_rain['time'], y=df_rain['intensity'])
            df_rain['intensity'].plot.bar()
            # x_label = df_rain['time'].values
            # print(x_label)
            # plt.xlabel()
            plt.savefig(Path.cwd()/f'precipitations/usage/charts/降雨強度_usage_{year}{month:02}{day:02}.png', dpi=300)
            # plt.show()
            plt.clf()
            plt.close()

def make_rainfall_bar_1hour(year, month, day, start_end_tuple:tuple): # 卒論用に1時間降雨のグラフを作成する。
    url = f'https://www.data.jma.go.jp/obd/stats/etrn/view/hourly_s1.php?prec_no=48&block_no=47618&year={year}&month={month}&day={day}&view='
    table = pd.read_html(url)
    df = table[0]
    # print(df)
    
    if start_end_tuple[0] == 0:
        start_end_tuple = (1, start_end_tuple[1])
    
    df.replace('--', '0.0', inplace=True)
    # print(df)
    df = df.astype({('降水量 (mm)','降水量 (mm)'): float})
    df = df.iloc[(start_end_tuple[0]-1):(start_end_tuple[1]), :]
    df = df.reset_index(drop=True)
    df_rain = df['降水量 (mm)']['降水量 (mm)']
    df_time = pd.Series([t for t in range(start_end_tuple[0], start_end_tuple[1]+1)], name='time') # 時刻
    
    # print(df_rain)
    # print(df_time)
    
    df_rain = pd.concat([df_time, df_rain], axis=1)
    df_rain = df_rain.set_index('time', drop=True)
    # print(df_rain)
    
    df_rain['降水量 (mm)'].plot.bar()
    plt.xticks(rotation=0)
    # plt.xlim(14, 21)
    plt.ylim(0.0, 15.5)
    plt.xlabel('time(時)')
    plt.ylabel('降水量\n(mm)', labelpad = 15, rotation = "horizontal")
    # plt.legend()
    # plt.show()
    plt.savefig(Path.cwd().parent.parent / f'grad_thesis/charts/precipitations/{year}{month:02}{day:02}.png', dpi=300)
    plt.clf()
    plt.close()
    
def make_rainfall_bar_1hour_extra(year, month, day1, day2, start_end_tuple:tuple): # 卒論用に1時間降雨のグラフを作成する。
    url = f'https://www.data.jma.go.jp/obd/stats/etrn/view/hourly_s1.php?prec_no=48&block_no=47618&year={year}&month={month}&day={day1}&view='
    url2 = f'https://www.data.jma.go.jp/obd/stats/etrn/view/hourly_s1.php?prec_no=48&block_no=47618&year={year}&month={month}&day={day2}&view='
    table1 = pd.read_html(url)
    table2 = pd.read_html(url2)
    df = table1[0]
    df2 = table2[0]
    # print(df)
    # print(df2)
    
    df = pd.concat([df, df2], axis=0)
    # print(df)
    
    df.replace('--', '0.0', inplace=True)
    
    df = df.astype({('降水量 (mm)','降水量 (mm)'): float})
    df = df.reset_index(drop=True)
    # print(df)
    df = df.iloc[(start_end_tuple[0]-1):(start_end_tuple[1] + 23), :]
    # print(df)
    
    df_rain = df['降水量 (mm)']['降水量 (mm)']
    # df_time = pd.Series([t for t in range(start_end_tuple[0], start_end_tuple[1]+1)], name='time') # 時刻
    df_time = df['時']['時']
    
    # print(df_rain)
    # print(df_time)
    
    df_rain = pd.concat([df_time, df_rain], axis=1)
    df_rain.rename(columns={'時':'time'}, inplace=True)
    df_rain = df_rain.set_index('time', drop=True)
    # print(df_rain)
    
    df_rain['降水量 (mm)'].plot.bar()
    plt.xticks(rotation=0)
    # plt.xlim(14, 21)
    plt.ylim(0.0, 15.5)
    plt.xlabel('time(時)')
    plt.ylabel('降水量\n(mm)', labelpad = 15, rotation = "horizontal")
    # plt.legend()
    # plt.show()
    plt.savefig(Path.cwd().parent.parent / f'grad_thesis/charts/precipitations/{year}{month:02}{day1:02}-{day2:02}.png', dpi=300)
    # plt.clf()
    # plt.close()
    
def make_extra_rain_df(year, month, day1, day2, start_end_tuple):
    RAINFALL_XLSX = Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day1:02}.xlsx'
    RAINFALL_XLSX2 = Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day2:02}.xlsx'
    rain_lis = []
    time_lis = []
    
    start_hour = start_end_tuple[0]
    end_hour = start_end_tuple[1]
    if 1 <= (start_hour * 60 / 10) <= 143:
        start_data_num = int(start_hour * 60 / 10) # 00:10から数えた順番
    elif start_hour==0:
        start_data_num = 1
    else:
        print(f'start_data_num:{(start_hour+1) * 60 / 10}は0から143の間でなければならない')
        
    if 2<= (end_hour * 60 / 10) <= 144:
        end_data_num = int(end_hour * 60 / 10)
    else:
        (f'end_data_num:{end_hour * 60 / 10}は1から144の間でなければならない')
    
    
    wb:openpyxl.Workbook = openpyxl.load_workbook(RAINFALL_XLSX)
    ws = wb.worksheets[0]
    
    wb2:openpyxl.Workbook = openpyxl.load_workbook(RAINFALL_XLSX2)
    ws2 = wb2.worksheets[0]
    
    
    count = 0
    for num in range(start_data_num, (144+1)):
        if (col_num := num%8) ==0:
            col_num = 8
        # else:
        #     col_num=col_num
        
        rain_lis.append(ws.cell(row=(((num - col_num)/8 + 1) +2), column=col_num).value)
        time_lis.append(count*10)
        count += 1
        
    for num in range(1, end_data_num+1):
        if (col_num := num%8) ==0:
            col_num = 8
        # else:
        #     col_num=col_num
        
        rain_lis.append(ws2.cell(row=(((num - col_num)/8 + 1) +2), column=col_num).value)
        time_lis.append(count*10)
        count += 1
        
        
        
        
    rain_array = np.array(rain_lis)
    minute_array = np.array(time_lis)
    
    array_2d = np.stack([minute_array, rain_array], 1)
    # print(array_2d)
    df_rain = pd.DataFrame(array_2d, columns=['time', 'intensity'])
    # print(df_rain)
    RAIN_DF_PATH = Path.cwd() / f'precipitations/usage/rainfall(intensity)_usage_{year}{month:02}{day1:02}-{day2:02}.csv'
    
    df_rain.to_csv(RAIN_DF_PATH, header=True, index=False)



if __name__ == '__main__':
    # format_rainfall(2015, 6, 26)
    
    # rainfallファイルを月一括で作る。
    # year = 2017
    # month = 10
    # day_lis = [x for x in range(1, 31)]
    # for day in day_lis:
    #     format_rainfall(year, month, day)
    
    # 時間定義ファイルは個別で作る。降雨データの開始時刻と終了時刻を設定するため。
    
    year = 2019
    
    # 2月
    # month = 2
    # day_lis = [19]
    # time_tuple_lis = [(14,20)]
    
    # 3月 6 19:00 -7 09:00, 10 19:00 -11 13:00
    # month = 3
    # day_lis =        [3     ,6      ,10     ,28]
    # time_tuple_lis = [(0,22),(21,24),(20,24),(3,8)]
    
    # 4月
    # month = 4
    # day_lis =        [10    ,26    ,27   ,30]
    # time_tuple_lis = [(4,23),(7,12),(0,5),(1,9)]
    
    
    # 5月
    # month = 5
    # day_lis =        [  6    , 21,   ]
    # time_tuple_lis = [(17,23), (1,11)]
    
    # 6月
    # month = 6
    # day_lis =        [  7    , 9   ,  10     ,  11      ,15     ,28   , 30]
    # time_tuple_lis = [(8,16),(16,24), (11, 19), (17, 23),(1, 10),(3, 9), (0, 10)]
    
    # # 7月
    # month =7
    # day_lis =        [4      ,12    ,14     ,17     ,18     ,22     ,27     ,29     ]
    # time_tuple_lis = [(0, 10),(0, 6),(16,24),(17,22),(11,24),(11,17),(17,24),(16,21)]
    
    # 8月
    # month =8
    # day_lis =        [3      ,6      ,11     ,16    ,19     ,20     ,21      ,22    ,27      ,29   ,30]
    # time_tuple_lis = [(14,19),(15,21),(15,21),(0,11),(15,24),(10,16),(14, 21),(6,12),(17, 24),(1,7),(8,23)]
    
    # 9月
    # month =9
    # day_lis =        [4      ,11     ,29]
    # time_tuple_lis = [(14,21),(13,22),(15,20)]
    
    # 10月　# 18 14:00-19 10:00, 21 20:00 -22 9:00, 24 20:00 -25 20:00は別途対応
    # month =10
    # day_lis =        [4    ,12    ,14   ,19     ]
    # time_tuple_lis = [(0,7),(0,24),(4,8),(13,19)]
    
    # 12月
    # month =12
    # day_lis =        [2]
    # time_tuple_lis = [(6, 20)]
    
   
    
    ## main ##
    # for day, start_end in zip(day_lis, time_tuple_lis):
    #     format_rainfall(year, month, day)
    #     format_timedef(year, month, day, start_end, make_df=True)
    
    
    # 対象とした降雨日の日付4桁を生成
    # month_lis = [2, 3, 4, 5, 6, 7, 8, 9, 10, 12]
    
    # day_lis2d = [[19],
    #              [4     ,6      ,10     ,28],
    #              [10    ,26    ,27   ,30],
    #              [  6    , 21,   ],
    #              [  7    , 9   ,  10     ,  11      ,15     ,28   , 30],
    #              [4      ,12    ,14     ,17     ,18     ,22     ,27     ,29     ],
    #              [3      ,6      ,11     ,16    ,19     ,20     ,21      ,22    ,27      ,29   ,30],
    #              [4      ,11     ,29],
    #              [4    ,12    ,14   ,19     ],
    #              [2]]
    
    # time_tuple_lis2d = [[(14,20)],
    #                     [(0,22),(21,24),(20,24),(3,8)],
    #                     [(4,23),(7,12),(0,5),(1,9)],
    #                     [(17,23), (1,14)],
    #                     [(8,16),(16,24), (11, 19), (17, 23),(1, 10),(3, 9), (0, 13)],
    #                     [(0, 10),(0, 6),(16,24),(17,22),(11,24),(11,17),(17,24),(16,21)],
    #                     [(14,19),(15,21),(15,21),(0,11),(15,24),(10,16),(14, 21),(6,12),(17, 24),(1,7),(8,23)],
    #                     [(14,21),(13,22),(15,20)],
    #                     [(0,7),(0,24),(7,20),(13,19)],
    #                     [(6, 20)]]
    
    # 使わない(もしグラフを変更するなら要変更)
    # 2/19 -> (13,21), 9/29 -> (14, 21)
    # 3/10 -> (18, 24), 3/28 -> (1,8), 4/26 -> (4,12), 5/6 -> (16,23)
    # 6/7 -> (6,18), 
    # month_lis = [3]
    # day_lis2d = [[4]]
    # time_tuple_lis2d = [[(0,24)]]
    
    # for i, month in enumerate(month_lis):
    #     day_lis = day_lis2d[i]
    #     time_tuple_lis = time_tuple_lis2d[i]
    #     # print(time_tuple_lis)
    #     for day, start_end in zip(day_lis, time_tuple_lis):
    #         format_rainfall(year, month, day)
    #         format_timedef(year, month, day, start_end, make_df=True, make_graph=True)
    #         # make_rainfall_bar_1hour(year, month, day, start_end)
    #         # print(year, month, day, start_end)
    

    
    
    # 日付跨っている文
    # format_rainfall_extra(2019, 10, 21, 22)
    
    make_rainfall_bar_1hour_extra(2019, 10, 21, 22, (20, 13))
    
    
    # make_extra_rain_df(2019, 3, 6, 7, (19, 9))
    # make_extra_rain_df(2019, 3, 10, 11, (19, 13))
    # make_extra_rain_df(2019, 10, 18, 19, (14, 10))
    # make_extra_rain_df(2019, 10, 21, 22, (20, 9))
    # make_extra_rain_df(2019, 10, 24, 25, (20, 20))