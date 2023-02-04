import pandas as pd
import openpyxl
from pathlib import Path
from decimal import Decimal

import numpy as np

# 閲覧用エクセルファイルを、NILIMフォーマットに最終変換する。
KANRO_XLSX = Path.cwd() / r'inputfiles_edit/KANRO.xlsx'
KANRO_FORMAT = Path.cwd().parent / r'inputfiles/KANRO.csv'

MANHOLE_XLSX = Path.cwd() / r'inputfiles_edit/MANHOLE.xlsx'
MANHOLE_FORMAT = Path.cwd().parent / r'inputfiles/MANHOLE.csv'

MESH_XLSX = Path.cwd() / r'inputfiles_edit/MESH.xlsx'
MESH_FORMAT = Path.cwd().parent / r'inputfiles/mesh.txt'


# 管路データ
kanro_xlsx = pd.read_excel(KANRO_XLSX)
kanro_xlsx.to_csv(KANRO_FORMAT, index=None, header=False)

# 人孔データ
man_xlsx = pd.read_excel(MANHOLE_XLSX)
man_xlsx.to_csv(MANHOLE_FORMAT, index=None, header=False)


# メッシュデータ
f_txt = open(MESH_FORMAT, 'w', newline='')

wb:openpyxl.Workbook = openpyxl.load_workbook(MESH_XLSX)
ws = wb.worksheets[0]

# 1行目
a2 = ws['A2'].value
b2 = ws['B2'].value
c2 = ws['C2'].value

first_row_lis = [f'{a2:>5}', f'{b2:>10.3f}', f'{c2:>10.3f}\n']
first_string = ''.join(first_row_lis)
f_txt.write(str(first_string))

# 2行目以降
test_lis = []
all_lis = []
for row in ws.iter_rows(min_row=4):
    temp_lis = []
    test_temp=[]
    for i, cell in enumerate(row):
        st = Decimal(cell.value)
        # print((i, st))
        test_temp.append(str(st))

    test_lis.append(test_temp)
    
for line in test_lis:
    
    n_row = list(map(Decimal, line))# f文字列に食べさせるために一度str型からDecimal型に変換
    
    temp_lis = []
    for i in range(0, 2):
        temp_lis.append(f'{n_row[i]:>5}')
    for i in range(2, 6):
        temp_lis.append(f'{n_row[i]:>10}')
    for i in range(6, 7):
        temp_lis.append(f'{n_row[i]:>10.2f}')
    for i in range(7, 8):
        temp_lis.append(f'{n_row[i]:>10.4f}')
    for i in range(8, 10):
        temp_lis.append(f'{n_row[i]:>10.2f}')    
    for i in range(10, 14):
        temp_lis.append(f'{n_row[i]:>10}')
    for i in range(14, 15):
        temp_lis.append(f'{n_row[i]:>10}\n')
    n_string = "".join(temp_lis)
    # print(n_string)
    
    f_txt.write(str(n_string))

f_txt.close()
        