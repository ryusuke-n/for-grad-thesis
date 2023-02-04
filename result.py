from PublicClass import Kanro

import openpyxl
from pathlib import Path
# 結果処理・表示のためのファイル

if __name__ == '__main__':
    # testcase = '20191014'
    # df_testcase = Kanro.txt_to_dataframe(testcase)
    # df_407 = Kanro.get_my_result(df_testcase, 1, 407)
    # # Kanro.display_info(df_407, 1, 407)
    # df_407_levels = Kanro.get_waterlevels(df_407, display=True)
    # # print(df_407)
    # Kanro.display_graph_hdw(df_407_levels, pipe_num = 407, case_date=testcase)
    # Kanro.display_graph_qdw(df_407_levels, pipe_num = 407, case_date=testcase)
    
    # registrate_lis = [376, 377, 378, 379, 380, 381, 382, 383, 384, 385, 386, 387, 388, 389, 390, 391, 392, 406, 407,\
    #                   120, 148, 173, 174, 171, 176, 177, 178, 184, 185, 193, 192, 190, 188, 198, 195, 194, \
    #                   201, 200, 199, 210, 208, 211, 213, 215, 206, 207, 205, 203, 202, 216, 217, 218, 219, 223, 224,\
    #                   235, 234, 233, 236, 237, 238, 405, 404, 403, 328, 325, 322, 321, 320, 319, 220, 221, 222,\
    #                   345, 343, 341]
    
    # 2, 3,4,
    # [19],
    #              [3     ,6      ,10     ,28],
    # [10    ,26    ,27   ,30],
    # ,21 #(8月)
    
    ##analysis_accumulationファイルの作成 ##
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = 'all_results'
    # ws['A1'] = '日付'
    # ws['B1'] = '最大1時間降水量(mm)'
    # ws['C1'] = '時間内総雨量(mm)'
    # ws['D1'] = '最大流量(mm/s)'
    # ws['E1'] = '総流量(m3)'
    # ws['F1'] = '越流の発生'
    # ws['G1'] = '越流時間(min)'
    
    # wb.save(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
    ####
    
    
    registrate_lis = [405]
    
    k = 1
    # year = 2019
#     month_lis = [  2, 3, 4, 5, 6, 7, 8, 9, 10, 12]
#     day_lis2d = [[19],
#                  [28],
#                  [10,    26, 27, 30],
#                  [  6 ,21   ],
#                  [  7    , 9   ,  10     ,  11      ,15     ,28   , 30],
#                  [4      ,12    ,14     ,17     ,18     ,22     ,27     ,29     ],
#                  [3      ,6      ,11     ,16    ,19     ,20 ,    21      ,22    ,27      ,29   ,30],
#                  [4      ,11     ,29],
#                  [4    ,12    ,14   ,19     ],
#                  [2]]
   
#    # 最後に追加すること
#     #3/4
   
#     # 単独テスト用    
    # month_lis = [3]
    # day_lis2d = [[4]]
    
    # 対照実験用
    year = 2024
    month_lis = [2]
    day_lis2d = [[2, 4, 6, 8, 10, 12, 14, 16]]
    
    # # main続き
    for month, day_lis in zip(month_lis, day_lis2d):
        for day in day_lis:
            testcase = f'{year}{month:02}{day:02}'
            for kanro_num in registrate_lis:
                df_testcase = Kanro.txt_to_dataframe(testcase)
                df_kanro = Kanro.get_my_result(df_testcase, 1, kanro_num)
                df_kanro_levels = Kanro.get_waterlevels(df_kanro, display=False)
                # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
                Kanro.display_graph_qdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False, get_max=True)
                # Kanro.analyze_with_rainfall(df_kanro_levels, year, month, day)
        
    # 単独テスト用    
    
    # year = 2019
    # month = 8
    # day = 27
    # testcase = '20190827'
    
    # result_accumulation
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = 'all_results'
    # ws['A1'] = '日付'
    # ws['B1'] = '最大1時間降水量(mm)'
    # ws['C1'] = '時間内総雨量(mm)'
    # ws['D1'] = '最大流量(mm/s)'
    # ws['E1'] = '総流量(m3)'
    # ws['F1'] = '越流の発生'
    # ws['G1'] = '越流時間(min)'
    
    # wb.save(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
    
    # for kanro_num in registrate_lis:
    #     df_testcase = Kanro.txt_to_dataframe(testcase)
    #     df_kanro = Kanro.get_my_result(df_testcase, 1, kanro_num)
    #     df_kanro_levels = Kanro.get_waterlevels(df_kanro, display=False)
    #     # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    #     Kanro.display_graph_qdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    #     # print(df_kanro_levels)
    #     Kanro.analyze_with_rainfall(df_kanro_levels, year, month, day)
    
    
    
    # 以下、extra
    
    # df_testcase = Kanro.txt_to_dataframe_extra('20190306-07')
    # df_kanro = Kanro.get_my_result_extra(df_testcase, 1, 407)
    # df_kanro_levels = Kanro.get_waterlevels_extra(df_kanro, display=False)
    # # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    # Kanro.display_graph_qdw(df_kanro_levels, pipe_num = 407, case_date='20190306-07', display=False)
    # # Kanro.analyze_with_rainfall_extra(df_kanro_levels, 2019, 3, 6, 7)
    
    # df_testcase = Kanro.txt_to_dataframe_extra('20190310-11')
    # df_kanro = Kanro.get_my_result_extra(df_testcase, 1, 407)
    # df_kanro_levels = Kanro.get_waterlevels_extra(df_kanro, display=False)
    # # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    # Kanro.display_graph_qdw(df_kanro_levels, pipe_num = 407, case_date='20190310-11', display=False)
    # # Kanro.analyze_with_rainfall_extra(df_kanro_levels, 2019, 3, 10, 11)
    
    # df_testcase = Kanro.txt_to_dataframe_extra('20191018-19')
    # df_kanro = Kanro.get_my_result_extra(df_testcase, 1, 407)
    # df_kanro_levels = Kanro.get_waterlevels_extra(df_kanro, display=False)
    # # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    # Kanro.display_graph_qdw(df_kanro_levels, pipe_num = 407, case_date='20191018-19', display=False)
    # # Kanro.analyze_with_rainfall_extra(df_kanro_levels, 2019, 10, 18, 19)
    
    # df_testcase = Kanro.txt_to_dataframe_extra('20191021-22')
    # df_kanro = Kanro.get_my_result_extra(df_testcase, 1, 407)
    # df_kanro_levels = Kanro.get_waterlevels_extra(df_kanro, display=False)
    # # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    # Kanro.display_graph_qdw(df_kanro_levels, pipe_num = 407, case_date='20191021-22', display=False)
    # # Kanro.analyze_with_rainfall_extra(df_kanro_levels, 2019, 10, 21, 22)
    
    # df_testcase = Kanro.txt_to_dataframe_extra('20191024-25')
    # df_kanro = Kanro.get_my_result_extra(df_testcase, 1, 407)
    # df_kanro_levels = Kanro.get_waterlevels_extra(df_kanro, display=False)
    # # Kanro.display_graph_hdw(df_kanro_levels, pipe_num = kanro_num, case_date=testcase, display=False)
    # Kanro.display_graph_qdw(df_kanro_levels, pipe_num = 407, case_date='20191024-25', display=False)
    # # Kanro.analyze_with_rainfall_extra(df_kanro_levels, 2019, 10, 24, 25)