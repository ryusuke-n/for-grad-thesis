import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib
import os

import numpy as np
import openpyxl
from pathlib import Path
import csv
import re
import string
import math
import datetime

import networkx as nx

# 他のファイル
import mesh_protocols

# 共通定数
KANRO_RESULT = r'../results/qh_kan.dat'
# GRAPHSAVE_FOLDER = Path.cwd().parent/ 'results/{}/charts'

MATSUMOTO_DATABASE = r'database/gouryuu_kuiki.xlsx' # 松本市からいただいた「合流区域」のファイル、のコピー
COORDINATES_FILE = Path.cwd() / r'meshcoordinate/mesh_coordinates.xlsx'


class Kanro:
    database_lis:list # クラス変数の型のみ宣言。(def)database_to_listにて代入する。
    kanro_instances:dict[int, "Kanro"]
    usage_data_lis:list
    
    lowest_pipe_num = 407
    
    k_numset = {1}# 排水区の番号の設定。今回の研究では排水区は1のみなので固定。
    
    TEMP_XLSX = Path.cwd() / r'inputfiles_edit/KANRO.xlsx' # NILIMフォーマットに変換する前のファイル。
    
    KANRO_RESULT = '../results/{}/qh_kan.dat'
    
    def __init__(self, key_k:int, key_pipe:int) -> None:
        self.k_num = key_k
        if isinstance(key_pipe, int):
            self.pipe_num:int = key_pipe
        else:
            raise TypeError("引数key_pipe のデータ型が不正です。")
        for pipe in Kanro.database_lis:
            if pipe[0] == self.pipe_num:
                self.inflow_numset:set = pipe[1]
                self.outflow_numset:set = pipe[2]
                self.ie_up = float(pipe[3]) # 管底高（上流側）
                self.ie_bottom = float(pipe[4]) # 管底高（下流側）
                
                # 管路の形状タイプと幅・高さ
                s = pipe[5].replace('\s', '') # 空白を削除
                which_shape = re.match('[^0-9]+', s).group() # 形状の○、×、その他記法を抽出。
                # print(which_shape)
                diameter = s.replace(str(which_shape), '')
                # print(diameter)
                self.width = int(diameter)/1000 # (mm)->(m)
                self.height = int(diameter)/1000 # (mm)->(m)
                if which_shape == '○':
                    self.shape = 1
                elif which_shape == '□':
                    self.shape = 2
                else:
                    self.shape = 0 # その他にも記号があるが、基本的に使われない管路のはず（最後川に接続している管路など）なので、ひとまず円形とする。
                
                self.length = pipe[6] # 管路延長(m)
                self.roughness = 0.013 # 粗度係数。https://isabou.net/soft/petit/common/waterway/flowcalc/ReferRoughCoef.asp
                self.weir_height = 0.0 # 分流堰はなし。
                self.weir_width = 0.0
                self.territory = float(pipe[7]) * 10000 # 「排水面積(m2)」。「集水面積」としてマンホールに渡す。
                
                break
            
        self.neighbor_man_numset = set() # 接続するマンホールの番号を格納するための空集合
        self.upstream_mannum:int # 上流側のマンホールの番号を格納
        self.downstream_mannum:int # 下流側のマンホールの番号を格納
        
        self.coord_upstream:list[float, float] # 上流端のx,y座標
        self.coord_downstream:list[float, float]
                
        # self.time_col = df_kanro.loc[:, 'time']
        # self.height_dw_col = df_kanro.loc[:, 'h-dw(m)']
    
    def __str__(self) -> str:
        return f"管路番号-> {self.pipe_num}, 上流管路-> {self.inflow_numset}, 下流管路->{self.outflow_numset}"
    
    def __repr__(self) -> str: # リスト内のインスタンスを正常に表示するための呪文
        return self.__str__()

    ####モデル作成時####
    @classmethod # 全体を通して実行回数は1度のみの想定。
    def database_to_list(cls): # 松本市のエクセルデータから情報を抽出する。NILIMフォーマットに必要な項目をリストにする。
        filepath = Path.cwd() / MATSUMOTO_DATABASE
        wb = openpyxl.load_workbook(filepath)
        ws = wb['南深志第三(p83～114)']
        kanro_lis = []
        for row in ws['A11':'A1335']:
            for cell in row:
                if not all(is_empty2(cell) for cell in row):
                    if type(cell.value) == int: # この2行で、管路番号の入っているセルのみを抽出した。
                        outflow_set = set() # 集合
                        inflow_set = set()
                        for i in [-1, 0, 1]:
                            outflow_position = 'C' + str(cell.row + i) # cell.rowで現在の行番号を取得。C列と組み合わせて座標とする。                        
                            if not is_empty2(ws[outflow_position]) and (isinstance(ws[outflow_position].value, int)):
                                outflow = ws[outflow_position].value # 流出管番号を取得
                                outflow_set.add(outflow)
                            inflow_position = 'B' + str(cell.row + i)
                            if not is_empty2(ws[inflow_position]) and (isinstance(ws[inflow_position].value, int)):# セルの中身が空欄でないかつ、int型であるときに...
                                inflow = ws[inflow_position].value
                                inflow_set.add(inflow)
                        invert_elevation_up = ws['P' + str(cell.row)].value # 上流側の管底高
                        invert_elevation_bottom = ws['P' + str(cell.row + 1)].value # 下流側の管底高
                        diameter_with_its_shape = ws['L' + str(cell.row)].value # 内径と管路の形の情報「○250」のような。
                        length = ws['F' + str(cell.row)].value # 管路延長
                        territory = ws['D'+ str(cell.row)].value # 排水面積
                        kanro_lis.append([cell.value, inflow_set, outflow_set, invert_elevation_up, invert_elevation_bottom, diameter_with_its_shape, length, territory])
        cls.database_lis = kanro_lis # クラス変数に代入。
    
    @classmethod
    def narrow_down_registration(cls, inputlis:list):# 管路リストに載っているもので、databaseの中身を絞り込む。
        cls.usage_data_lis = []
        for pipe_num in inputlis:
            for data in cls.database_lis:
                if pipe_num == data[0]:
                    cls.usage_data_lis.append(data)
        # print(cls.usage_data_lis)
        # print(len(cls.usage_data_lis) == len(inputlis)) # 検索が全て成功しているかの判定
    
    @classmethod
    def cross_reference(cls):# 松本市の管路データベースには、上流・下流管路が相互に情報を持っていないため。
        for data in cls.usage_data_lis:
            id = data[0]
            self_inflowset = data[1]
            self_outflowset = data[2]
            for data2 in cls.usage_data_lis:
                if data2[0] in self_inflowset:
                    data2[2].add(id)
                if data2[0] in self_outflowset:
                    data2[1].add(id)
    
    @classmethod
    def create_kanro_instances(cls, k_num): # 管路インスタンスの生成して、クラス変数（辞書）に代入する。
        cls.kanro_instances = {} # dict型
        for p_num in cls.usage_data_lis:
            kanro = cls(k_num, p_num[0])
            cls.kanro_instances[p_num[0]] = kanro
            # print(kanro)
    
    @classmethod
    def start_make_connections(cls): # 管路同士を繋ぎ、マンホールインスタンスを生成する。make_connection_firstを起動
        lowest_pipe = cls.kanro_instances[cls.lowest_pipe_num]
        lowest_pipe.make_connection_first()
    
    def make_connection_first(self): # 最下流の管路インスタンスがはじめに一度だけ実行するメソッド。
        children_set = self.inflow_numset
        for child_num in children_set:
            child = Kanro.kanro_instances[child_num]
            child.make_connections(self.pipe_num)
        Manhole.manhole_instances[self.pipe_num] = Manhole(self.pipe_num, children_set)
        for child_num in children_set:
            pipe = Kanro.kanro_instances[child_num] # 子管路のインスタンスにアクセス。
            pipe.neighbor_man_numset.add(self.pipe_num) # このself.pipe_numは人孔番号としての意味。そのインスタンスの隣接人孔番号セットに追加する。
        self.neighbor_man_numset.add(self.pipe_num)
        
        Manhole.manhole_instances[0] = Manhole(0, {self.pipe_num}) # 最下流の人孔番号は、特別に0とする。
        Manhole.manhole_instances[0].neighbor_pipenum_set.remove(0) # 最下流の人孔の隣には、「管路0番」というものは存在しないので削除
        self.neighbor_man_numset.add(0)
    
    def make_connections(self, parent_num): # 再帰関数。n>=2において実施。
        neighbor_union = self.inflow_numset | self.outflow_numset # 和集合　# 「隣接する管路」の情報
        children_set = neighbor_union - {parent_num} # 指示された元の管路の番号を削除。# 子どもたちの番号
        remove_set = set()
        for child_num in children_set:
            try:
                child = Kanro.kanro_instances[child_num]
                child.make_connections(self.pipe_num)
            except KeyError:# もし今回対象としていない管路（registrate_lisに存在しない管路）があれば、try内容は無視して、子管路の集合から取り除く。
                remove_set.add(child_num)
        children_set = children_set - remove_set
        Manhole.manhole_instances[self.pipe_num] = Manhole(self.pipe_num, children_set) # マンホールインスタンスを生成して、一覧に追加。人孔番号は、親管路の管路番号と同じにする。
        # 管路側にも人孔番号の情報を与える。
        for child_num in children_set:
            pipe = Kanro.kanro_instances[child_num] # 子管路のインスタンスにアクセス。
            pipe.neighbor_man_numset.add(self.pipe_num) # このself.pipe_numは人孔番号としての意味。そのインスタンスの隣接人孔番号セットに追加する。
        self.neighbor_man_numset.add(self.pipe_num)
    
    @classmethod
    def start_discrimination(cls): # 隣接人孔を上流側と下流側で区別。
        for kanro in cls.kanro_instances.values():
            kanro.discriminate_manholes()
            # print(f"管路番号{kanro.pipe_num}| 上流-> {kanro.upstream_mannum} 下流-> {kanro.downstream_mannum}")
        
    def discriminate_manholes(self): # 隣接するマンホールが、上流側なのか下流側なのかを判別する。
        for man_num in self.neighbor_man_numset:
            # print(self.neighbor_man_numset)
            manhole = Manhole.manhole_instances[man_num]
            # print(f"管路番号->{self.pipe_num} 人孔番号->{manhole.man_id} 人孔に隣接する管路->{manhole.neighbor_pipenum_set}")
            if inflow_num := self.inflow_numset & manhole.neighbor_pipenum_set:# もし、管路の「流入管一覧」とマンホールの「隣接する管路一覧」に被りがあるなら、このmanholeは上流側になる。
                self.upstream_mannum = man_num
                # print(self.pipe_num, self.inflow_numset, inflow_mannum_set)  
            elif outflow_num := self.outflow_numset & manhole.neighbor_pipenum_set:
                self.downstream_mannum = man_num
                # print(f"管路番号->{self.pipe_num} 人孔番号->{manhole.man_id} {outflow_num}")
            elif (inflow_num == set()) and (man_num != 0): # 最上流の管路の場合、管路から見た「流入管路番号」と人孔から見た「隣接管路」の積集合は空集合となるはずである。
                self.upstream_mannum = man_num
                # print(f"管路番号->{self.pipe_num} 人孔番号->{manhole.man_id} {inflow_num}")
            elif (outflow_num ==set()) and (man_num == 0) : # 要修正
                self.downstream_mannum = 0
                # print(f"管路番号->{self.pipe_num} 人孔番号->{manhole.man_id}")
            else:
                print(f"通過しています。　番号->{man_num} by discriminate_manholes()")
    
    @classmethod
    def get_coordinates(cls): # 管路上下端の座標をエクセルデータより取得する。
        wb:openpyxl.Workbook = openpyxl.load_workbook(COORDINATES_FILE)
        ws = wb.worksheets[0]
        for row in ws.iter_rows():
            pipe_num = row[0].value
            try:
                pipe = cls.kanro_instances[pipe_num]
                pipe.coord_upstream = [row[1].value, row[2].value]
                pipe.coord_downstream = [row[3].value, row[4].value]
            except KeyError:
                print(f"{pipe_num}番は該当なし。除外されました。")
    
    @classmethod
    def start_give_coordinates(cls):
        for kanro in cls.kanro_instances.values():
            kanro.give_coordinates_to_manhole()
        # 最下流のマンホールの操作
        kanro0 = cls.kanro_instances[cls.lowest_pipe_num]
        manhole0 = Manhole.manhole_instances[kanro0.downstream_mannum]
        manhole0.coordinates_lis = kanro0.coord_downstream
        
        
    
    def give_coordinates_to_manhole(self):
        try:
            manhole = Manhole.manhole_instances[self.upstream_mannum] # 上流のマンホールを呼び出す。
        except AttributeError:
            print(self.pipe_num)
        try:
            manhole.coordinates_lis = self.coord_upstream
        except AttributeError:
            manhole.coordinates_lis = [None, None]
            print(f"座標を取得していません-> {self.pipe_num}番")
        # print(f"人孔番号->{manhole.man_id} 座標->{manhole.coordinates_lis}")
    
    @classmethod
    def to_networkx(cls):
        gnl = [] # 頂点のリスト
        gel = [] # 辺のリスト
        gpos = {} # 頂点と座標の辞書
        e_label_dict = {} # 頂点の名前
        
        escape_count = 0
        for manhole in Manhole.manhole_instances.values():
            gnl.append(manhole.man_id)
            if manhole.coordinates_lis == [None, None]:
                gpos[manhole.man_id] = (-10-escape_count, -10-escape_count)
                escape_count += 100
            else:
                gpos[manhole.man_id] = tuple(manhole.coordinates_lis)
        for kanro in cls.kanro_instances.values():
            gel.append((kanro.upstream_mannum, kanro.downstream_mannum))
            e_label_dict[(kanro.upstream_mannum, kanro.downstream_mannum)] = kanro.pipe_num
        g = nx.DiGraph()
        g.add_nodes_from(gnl)
        g.add_edges_from(gel)
        
        
        # print('nodes of graph g:')
        # print(g.nodes())
        
        # print('edges of graph g:')
        # for edge in g.edges():
        #     print(edge)
        
        nx.draw_networkx(g, gpos, node_size=150, with_labels=True, font_size=6, node_color="#cccccc")
        nx.draw_networkx_edge_labels(g, gpos,edge_labels=e_label_dict, font_color='b', font_size=8)
        plt.show()
    
    def output_info(self):
        if self.upstream_mannum == 0:
            upstream_mannum = 1
        else:
            upstream_mannum = self.upstream_mannum
        if self.downstream_mannum == 0:
            downstream_mannum = 1
        else:
            downstream_mannum = self.downstream_mannum
        # 管路データ
        return np.array([self.pipe_num, upstream_mannum, downstream_mannum, self.shape, self.k_num, self.length, self.width, self.height, self.roughness, self.ie_up, self.ie_bottom, self.weir_height, self.weir_width])
        
    
    @classmethod
    def output_kanro_csvformat(cls): # データを閲覧・編集するためにcsvへ
        data_array = np.empty((0, 13))
        
        
        for i, kanro in enumerate(cls.kanro_instances.values()):
            data_array = np.append(data_array, np.array([kanro.output_info()]), axis=0)
          
        np.set_printoptions(suppress=True)
        # print(data_array)
        # np.savetxt(Path.cwd() / 'inputfiles_edit/KANRO.csv', data_array, delimiter=',')
        
        # 途中閲覧用のエクセルファイルを作る。
        # try:
        #     wb = openpyxl.load_workbook(cls.TEMP_XLSX)
        # except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
               
        header_lis = ['管路番号','上流人孔番号','下流人孔番号','管路形式','排水区番号','管路長','管路幅','管路高','粗度係数','上流管底高','下流管底高','分流堰高','分流堰幅']
        for i,header in enumerate(header_lis):
            ws.cell(row=1, column=i+1, value=header)
        
        write_list_2d(ws, data_array, start_row=2, start_col=1)
        
        wb.save(cls.TEMP_XLSX)
        
            

    ####NILIM2.0にて解析後、結果表示####
    @classmethod
    def txt_to_dataframe(cls, date): #全体を通して一度だけ実行する。 # txtの中身を空白で分けてリスト化し,さらにデータフレーム化したものを返す。# シミュレーション結果を表示する前に使う。
        filepath = cls.KANRO_RESULT.format(date)
        row_lis = []
        with open(filepath) as f:# txt形式からリストにする。
            for row in f:
                row_lis.append(row.split())
            for line in row_lis: # 一番右の列は不要なので削除
                if len(line)==23:
                    del line[22]
                else:
                    pass
        df_result = pd.DataFrame(row_lis, columns=row_lis[0])# データフレーム化。columnsオプションで列名を設定
        df_result.drop([0], axis=0, inplace=True)
        df_result.sort_values(by=['k', 'pipe'], inplace=True)# 排水区番号と管路番号で並び替え
        df_result.reset_index(inplace=True)
        # ここで各列のデータ型を変換
        df_result = df_result.astype({'time':int, 'k':int, 'pipe':int, 'man-up':float, 'man-dw':float, 'gl-up(m)':float, 'gl-dw(m)':float, 'btm-up(m)':float, 'btm-dw(m)':float, 'top-up(m)':float, 'top-dw(m)':float, 'q-up(m3/s)':float, 'q-dw(m3/s)':float, 'h-up(m)':float, 'h-dw(m)':float, 'hms-up(m)':float, 'hms-dw(m)':float, 's-up(m3/s)':float, 's-dw(m3/s)':float, 'line':int, 'length(m)':float, 'figure':float})
        return df_result
    
    @staticmethod
    def get_my_result(df_result: pd.DataFrame, k_num, pipe_num): # 自分（管路）自身のデータフレームだけを抽出して保持するメソッド。
        df_kanro = df_result[df_result['k'].isin([int(k_num)])] # 一致する排水区の行データを抽出
        df_kanro = df_kanro[df_kanro['pipe'].isin([int(pipe_num)])] # 一致する管路の行データを抽出
        return df_kanro
    
    @classmethod
    def get_waterlevels(cls, df_kanro, display=False, k_num=None, pipe_num=None):
        df_levels = df_kanro.loc[:, ['time', 'h-dw(m)','q-dw(m3/s)']]
        df_levels['time'] = df_levels['time'] / 60 # 時間を秒数表記から分数表記に変換する。
        df_levels['time'] = df_levels['time'].astype('int')
        # print(df_levels)
        if display == True:
            cls.display_info(df_levels, k_num, pipe_num)
        return df_levels
    
    @staticmethod
    def display_info(df_levels, k_num=None, pipe_num=None):
        # print(f"排水区 - 管路番号 => {k_num} - {pipe_num}")
        # print(df_kanro.loc[:, ['time', 'h-dw(m)']])
        print(df_levels)
        

    
    @staticmethod
    def display_graph_hdw(df_levels, pipe_num, case_date, col_lis = ['time', 'h-dw(m)'], display=True): # 水位のグラフを作成して保存・表示するメソッド
        df_graph = df_levels.loc[:, col_lis]
        
        # 以下、グラフ作成
        df_graph.plot(x='time', y='h-dw(m)', linewidth=4.0, color='#0096c8')
        plt.title(f"管路(No.{pipe_num})下端における水位時間変化-({case_date})", fontsize=15)
        plt.xlabel("時間（s）", fontsize=18)
        plt.ylabel("水位\n（m）", labelpad = 35, size = "xx-large", rotation = "horizontal")
        plt.xticks(fontsize=11)
        plt.yticks(fontsize=11)
        plt.ylim(0.00, 1.50)
        
        # plt.rcParams['figure.subplot.bottom'] = 0.20 # 見切れ防止
        plt.subplots_adjust(left=0.20, bottom=0.15)
        plt.tight_layout()
        
        filepath = Path.cwd().parent/ f'results/{case_date}/charts'
        
        if not os.path.exists(filepath):# 保存用フォルダ。なければここで作成
            os.makedirs(filepath)

        plt.savefig(filepath / f"Graph_{case_date}_pipe_{pipe_num}.png", dpi=300)
        if display == True:
            plt.show()
    
    @staticmethod
    def display_graph_qdw(df_levels, pipe_num, case_date, col_lis = ['time', 'q-dw(m3/s)'], display=True, get_max=False): # 流量の計算
        df_graph = df_levels.loc[:, col_lis]
        # time（分数）をリセットする。
        df_graph['time'] = df_graph['time'] - df_graph.iloc[0,0] # df_graph.iloc[0,0]は'time'の一行目
        df_graph = df_graph.reset_index(drop=True)
        
        # print(df_graph)
        
        # 以下、グラフ作成
        df_graph.plot(x='time', y='q-dw(m3/s)', linewidth=4.0, color='#0096c8')
        # plt.title(f"管路({pipe_num})下端における流量時間変化-({case_date})", fontsize=13)
        plt.xlabel("時間（min）", fontsize=12)
        plt.ylabel("流量\n（m3/s）", labelpad = 20, rotation = "horizontal", fontsize=12)
        plt.xticks(fontsize=11)
        plt.yticks(fontsize=11)
        plt.ylim(0.00, 2.00)
        
        # 遮集能力を表す補助線
        plt.hlines(y=0.368, xmin=df_graph['time'].min(), xmax=df_graph['time'].max(), colors='red',linestyle='dashed', linewidth=1.5)
        
        # plt.rcParams['figure.subplot.bottom'] = 0.20 # 見切れ防止
        plt.subplots_adjust(left=0.20, bottom=0.15)
        plt.tight_layout()
        
        filepath = Path.cwd().parent/ f'results/{case_date}/charts'
        
        if not os.path.exists(filepath):# 保存用フォルダ。なければここで作成
            os.makedirs(filepath)

        # plt.savefig(filepath / f"Graph_{case_date}_pipe_{pipe_num}_q.png", dpi=300)
        
        chart_path = Path.cwd().parent.parent / 'grad_thesis/charts/kan_q'
        plt.savefig(chart_path / f"{case_date}.png", dpi=300)
        
        
        if display == True:
            plt.show()
        
        plt.clf()
        plt.close()
        
        if get_max == True:
            print( 'casedate->'+str(case_date)+' | max->'+str(df_graph['q-dw(m3/s)'].max()))
    
    @staticmethod
    def analyze_with_rainfall(df_levels, year, month, day):
        # print(df_levels)
        case_date = f'{year}{month:02}{day:02}'
        
        # df_levels = df_levels['q-dw(m3/s)']
        df_levels = df_levels.reset_index(drop=True) # 後で横結合するために、インデックスをリセット
        # print(df_levels)
        
        RAIN_DF_PATH = Path.cwd() / f'precipitations/usage/rainfall(intensity)_usage_{case_date}.csv'
        
        df_rain = pd.read_csv(RAIN_DF_PATH)
        
        df_rain['precipitation'] = df_rain['intensity'] / 6
        df_rain = df_rain.reset_index(drop=True)
        # print(df_rain)
    
        df_result = pd.concat([df_rain,df_levels], axis='columns')
      
        # print(df_result)
        
        # print(df_result['q-dw(m3/s)'] > 0.368) # 駅前雨水吐きの満管遮集能力
        # print(df_result['q-dw(m3/s)'].max())
        # print(df_result['precipitation'].max())
        
        ANALYSIS_CSVPATH = Path.cwd().parent / f'analysis/analysis_{case_date}.csv'
        df_result.to_csv(ANALYSIS_CSVPATH, index=False, encoding='shift_jis')
        ANALYSIS_XLSXPATH = Path.cwd().parent / f'analysis/analysis_{case_date}.xlsx'
        df_result.to_excel(ANALYSIS_XLSXPATH, index=False)
        
        # analysis シート
        wb = openpyxl.load_workbook(ANALYSIS_XLSXPATH)
        ws = wb.create_sheet('analysis')
        
        ws['A1'] = '日付'
        ws['B1'] = '最大1時間降水量(mm)'
        ws['C1'] = '時間内総雨量(mm)'
        ws['D1'] = '最大流量(mm/s)'
        ws['E1'] = '総流量(m3)'
        ws['F1'] = '越流の発生'
        ws['G1'] = '越流時間(min)'
        
        # 日付
        ws['A2'] = datetime.date(year=year, month=month, day=day)
        
        # 最大1時間雨量
        df_dailymax = pd.read_csv(Path.cwd() / f'precipitations/monthly/dailymax/dailymax_{year}{month:02}.csv',index_col='日' ,encoding='shift_jis')
        ws['B2'] = float(df_dailymax.loc[day, '最大1時間雨量'])
        
        # 時間内総雨量
        ws['C2'] = df_result['precipitation'].sum()
        
        # 最大流量
        ws['D2'] = df_result['q-dw(m3/s)'].max()
        
        # 総流量
        df_result['10分間流量'] = df_result['q-dw(m3/s)'] * 600
        ws["E2"] = df_result['10分間流量'].sum()
        
        # 越流の発生・総時間
        df_result['overflow'] = (df_result['q-dw(m3/s)'] > 0.368)
        if (overflow_count := df_result['overflow'].sum()) >=1:
            ws['F2'] = 1 # 越流あり
            ws['G2'] = overflow_count * 10
        elif overflow_count == 0:
            ws['F2'] = 0 # 越流なし
            ws['G2'] = overflow_count * 10
        # print(df_result)
        df_result.to_csv(ANALYSIS_CSVPATH, index=False,encoding='shift_jis')
        df_result.to_excel(ANALYSIS_XLSXPATH, index=False)
        
        # データを蓄積する。
        wb2 = openpyxl.load_workbook(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
        ws2 = wb2['all_results']
        
        ws2.append([ws['A2'].value, ws['B2'].value, ws['C2'].value, ws['D2'].value, ws['E2'].value, ws['F2'].value, ws['G2'].value])
        wb2.save(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
        
        
        wb.save(ANALYSIS_XLSXPATH)

###extra（日を跨ぐ用）###
    @classmethod
    def txt_to_dataframe_extra(cls, date): #全体を通して一度だけ実行する。 # txtの中身を空白で分けてリスト化し,さらにデータフレーム化したものを返す。# シミュレーション結果を表示する前に使う。
        filepath = cls.KANRO_RESULT.format(date)
        row_lis = []
        with open(filepath) as f:# txt形式からリストにする。
            for row in f:
                row_lis.append(row.split())
            for line in row_lis: # 一番右の列は不要なので削除
                if len(line)==23:
                    del line[22]
                else:
                    pass
        df_result = pd.DataFrame(row_lis, columns=row_lis[0])# データフレーム化。columnsオプションで列名を設定
        df_result.drop([0], axis=0, inplace=True)
        df_result.sort_values(by=['k', 'pipe'], inplace=True)# 排水区番号と管路番号で並び替え
        df_result.reset_index(inplace=True)
        # ここで各列のデータ型を変換
        df_result = df_result.astype({'time':int, 'k':int, 'pipe':int, 'man-up':float, 'man-dw':float, 'gl-up(m)':float, 'gl-dw(m)':float, 'btm-up(m)':float, 'btm-dw(m)':float, 'top-up(m)':float, 'top-dw(m)':float, 'q-up(m3/s)':float, 'q-dw(m3/s)':float, 'h-up(m)':float, 'h-dw(m)':float, 'hms-up(m)':float, 'hms-dw(m)':float, 's-up(m3/s)':float, 's-dw(m3/s)':float, 'line':int, 'length(m)':float, 'figure':float})
        return df_result
    
    @staticmethod
    def get_my_result_extra(df_result: pd.DataFrame, k_num, pipe_num): # 自分（管路）自身のデータフレームだけを抽出して保持するメソッド。
        df_kanro = df_result[df_result['k'].isin([int(k_num)])] # 一致する排水区の行データを抽出
        df_kanro = df_kanro[df_kanro['pipe'].isin([int(pipe_num)])] # 一致する管路の行データを抽出
        return df_kanro
    
    @classmethod
    def get_waterlevels_extra(cls, df_kanro, display=False, k_num=None, pipe_num=None):
        df_levels = df_kanro.loc[:, ['time', 'h-dw(m)','q-dw(m3/s)']]
        df_levels['time'] = df_levels['time'] / 60 # 時間を秒数表記から分数表記に変換する。
        df_levels['time'] = df_levels['time'].astype('int')
        # print(df_levels)
        if display == True:
            cls.display_info(df_levels, k_num, pipe_num)
        return df_levels
    
    @staticmethod
    def display_info_extra(df_levels, k_num=None, pipe_num=None):
        # print(f"排水区 - 管路番号 => {k_num} - {pipe_num}")
        # print(df_kanro.loc[:, ['time', 'h-dw(m)']])
        print(df_levels)
        

    
    @staticmethod
    def display_graph_hdw_extra(df_levels, pipe_num, case_date, col_lis = ['time', 'h-dw(m)'], display=True): # 水位のグラフを作成して保存・表示するメソッド
        df_graph = df_levels.loc[:, col_lis]
        
        # 以下、グラフ作成
        df_graph.plot(x='time', y='h-dw(m)', linewidth=4.0, color='#0096c8')
        plt.title(f"管路(No.{pipe_num})下端における水位時間変化-({case_date})", fontsize=15)
        plt.xlabel("時間（s）", fontsize=18)
        plt.ylabel("水位\n（m）", labelpad = 35, size = "xx-large", rotation = "horizontal")
        plt.xticks(fontsize=11)
        plt.yticks(fontsize=11)
        plt.ylim(0.00, 1.50)
        
        # plt.rcParams['figure.subplot.bottom'] = 0.20 # 見切れ防止
        plt.subplots_adjust(left=0.20, bottom=0.15)
        plt.tight_layout()
        
        filepath = Path.cwd().parent/ f'results/{case_date}/charts'
        
        if not os.path.exists(filepath):# 保存用フォルダ。なければここで作成
            os.makedirs(filepath)

        plt.savefig(filepath / f"Graph_{case_date}_pipe_{pipe_num}.png", dpi=300)
        if display == True:
            plt.show()
    
    @staticmethod
    def display_graph_qdw_extra(df_levels, pipe_num, case_date, col_lis = ['time', 'q-dw(m3/s)'], display=True): # 流量の計算
        df_graph = df_levels.loc[:, col_lis]
        # time（分数）をリセットする。
        df_graph['time'] = df_graph['time'] - df_graph.iloc[0,0] # df_graph.iloc[0,0]は'time'の一行目
        df_graph = df_graph.reset_index(drop=True)
        
        # print(df_graph)
        
        # 以下、グラフ作成
        df_graph.plot(x='time', y='q-dw(m3/s)', linewidth=4.0, color='#0096c8')
        # plt.title(f"管路({pipe_num})下端における流量時間変化-({case_date})", fontsize=13)
        plt.xlabel("時間（min）", fontsize=12)
        plt.ylabel("流量\n（m3/s）", labelpad = 20, rotation = "horizontal", fontsize=12)
        plt.xticks(fontsize=11)
        plt.yticks(fontsize=11)
        plt.ylim(0.00, 2.00)
        
        plt.hlines(y=0.368, xmin=df_graph['time'].min(), xmax=df_graph['time'].max(), colors='red',linestyle='dashed', linewidth=1.5)
        
        # plt.rcParams['figure.subplot.bottom'] = 0.20 # 見切れ防止
        plt.subplots_adjust(left=0.20, bottom=0.15)
        plt.tight_layout()
        
        filepath = Path.cwd().parent/ f'results/{case_date}/charts'
        
        if not os.path.exists(filepath):# 保存用フォルダ。なければここで作成
            os.makedirs(filepath)

        plt.savefig(filepath / f"Graph_{case_date}_pipe_{pipe_num}_q.png", dpi=300)
        
        chart_path = Path.cwd().parent.parent / 'grad_thesis/charts/kan_q'
        plt.savefig(chart_path / f"{case_date}.png", dpi=300)
        
        
        if display == True:
            plt.show()
        
        plt.clf()
        plt.close()
    
    @staticmethod
    def analyze_with_rainfall_extra(df_levels, year, month, day1, day2):
        # print(df_levels)
        case_date = f'{year}{month:02}{day1:02}-{day2:02}'
        
        # df_levels = df_levels['q-dw(m3/s)']
        df_levels = df_levels.reset_index(drop=True) # 後で横結合するために、インデックスをリセット
        # print(df_levels)
        
        RAIN_DF_PATH = Path.cwd() / f'precipitations/usage/rainfall(intensity)_usage_{case_date}.csv'
        
        df_rain = pd.read_csv(RAIN_DF_PATH)
        
        df_rain['precipitation'] = df_rain['intensity'] / 6
        df_rain = df_rain.reset_index(drop=True)
        # print(df_rain)
    
        df_result = pd.concat([df_rain,df_levels], axis='columns')
      
        # print(df_result)
        
        # print(df_result['q-dw(m3/s)'] > 0.368) # 駅前雨水吐きの満管遮集能力
        # print(df_result['q-dw(m3/s)'].max())
        # print(df_result['precipitation'].max())
        
        ANALYSIS_CSVPATH = Path.cwd().parent / f'analysis/analysis_{case_date}.csv'
        df_result.to_csv(ANALYSIS_CSVPATH, index=False, encoding='shift_jis')
        ANALYSIS_XLSXPATH = Path.cwd().parent / f'analysis/analysis_{case_date}.xlsx'
        df_result.to_excel(ANALYSIS_XLSXPATH, index=False)
        
        # analysis シート
        wb = openpyxl.load_workbook(ANALYSIS_XLSXPATH)
        ws = wb.create_sheet('analysis')
        
        ws['A1'] = '日付'
        ws['B1'] = '最大1時間降水量(mm)'
        ws['C1'] = '時間内総雨量(mm)'
        ws['D1'] = '最大流量(mm/s)'
        ws['E1'] = '総流量(m3)'
        ws['F1'] = '越流の発生'
        ws['G1'] = '越流時間(min)'
        
        # 日付
        ws['A2'] = datetime.date(year=year, month=month, day=day1)
        
        # 最大1時間雨量
        df_dailymax = pd.read_csv(Path.cwd() / f'precipitations/monthly/dailymax/dailymax_{year}{month:02}.csv',index_col='日' ,encoding='shift_jis')
        dailymax = max(float(df_dailymax.loc[day1, '最大1時間雨量']), float(df_dailymax.loc[day2, '最大1時間雨量']))
        ws['B2'] = dailymax
        # 時間内総雨量
        ws['C2'] = df_result['precipitation'].sum()
        
        # 最大流量
        ws['D2'] = df_result['q-dw(m3/s)'].max()
        
        # 総流量
        df_result['10分間流量'] = df_result['q-dw(m3/s)'] * 600
        ws["E2"] = df_result['10分間流量'].sum()
        
        # 越流の発生・総時間
        df_result['overflow'] = (df_result['q-dw(m3/s)'] > 0.368)
        if (overflow_count := df_result['overflow'].sum()) >=1:
            ws['F2'] = 1 # 越流あり
            ws['G2'] = overflow_count * 10
        elif overflow_count == 0:
            ws['F2'] = 0 # 越流なし
            ws['G2'] = overflow_count * 10
        # print(df_result)
        df_result.to_csv(ANALYSIS_CSVPATH, index=False,encoding='shift_jis')
        df_result.to_excel(ANALYSIS_XLSXPATH, index=False)
        
        # データを蓄積する。
        wb2 = openpyxl.load_workbook(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
        ws2 = wb2['all_results']
        
        ws2.append([ws['A2'].value, ws['B2'].value, ws['C2'].value, ws['D2'].value, ws['E2'].value, ws['F2'].value, ws['G2'].value])
        wb2.save(Path.cwd().parent / f'analysis/analysis_accumulation.xlsx')
        
        
        wb.save(ANALYSIS_XLSXPATH)

class Manhole:
    manhole_instances:dict[int, "Manhole"] = {}
    
    TEMP_XLSX = Path.cwd() / r'inputfiles_edit/MANHOLE.xlsx'
    
    def __init__(self, parent_arm_num:int, children_arm_set:set) -> None:
        self.man_id:int = parent_arm_num
        self.neighbor_pipenum_set:set = {parent_arm_num} | children_arm_set # このマンホールに接続している管路の番号一覧
        self.coordinates_lis:list[float, float]
        
        self.mesh_position:tuple[int, int]
        if self.man_id == 0:
            self.exit_flag = 4
        else:
            self.exit_flag = 0 # 出口フラッグ
        self.k_num = list(Kanro.k_numset)[0]
        
        
        
    def __str__(self) -> str:
        return f"人孔番号-> {self.man_id}"
    
    def __repr__(self) -> str:
        return self.__str__()
    
    def which_Mesh_and_elevation(self):
        self.mesh_position = Mesh.classify_Manhole(self)
        try:
            self.elevation = Mesh.mesh_instances[self.mesh_position].dem # 同時に、標高データも取得
        except KeyError:
            print(self.man_id, self.mesh_position, '範囲外のメッシュが選択されています。')
            self.elevation = 99999
    @classmethod
    def start_which_Mesh_and_elevation(cls):
        for manhole in cls.manhole_instances.values():
            manhole.which_Mesh_and_elevation()
            # print(manhole.elevation)
            
    # def how_much_elevation(self): # マンホールのあるメッシュの標高を取得する
    #     mesh:"Mesh" = Mesh.mesh_instances[self.mesh_position]
    #     self.elevation = mesh.dem
    
    def get_my_territory(self):
        if self.man_id != 0:
            kanro = Kanro.kanro_instances[self.man_id]
            self.territory_area = kanro.territory
            return self.territory_area
        elif self.man_id == 0:
            self.territory_area = 0 # 最下流のマンホールの集水面積は0とした。
            return self.territory_area

    def catch_water_from(self):
        delta_lis = [[x,y] for x in range(-2, 3) for y in range(-2, 3) if [x,y] not in [[-2,-2], [-2,2],[2,2],[2,-2]]]
        for x, y in delta_lis:
            position = (int(self.mesh_position[0]+x), int(self.mesh_position[1]+y))
            try:
                mesh:'Mesh' = Mesh.mesh_instances[position]
                mesh.manhole_to_supply_water.add(self.man_id)
            except KeyError: # メッシュの端っこであれば何もしない。
                pass
    
    @classmethod
    def start_catch_water_from(cls):
        for manhole in cls.manhole_instances.values():
            manhole.catch_water_from()
                    
    def output_info(self):
        if self.man_id == 0:
            man_id = 1
        else:
            man_id = self.man_id
        # 人孔データ
        return np.array([man_id, 0.6**2*math.pi, self.elevation, Mesh.position_to_serial(self.mesh_position), self.exit_flag, self.k_num, self.get_my_territory(), 95.0, 0.01, 0.011, 1, 0, 0, 0, 0, 0, self.coordinates_lis[0], self.coordinates_lis[1]])
        # 
        # 断面積は、内径を600mmとして算出。
    
    @classmethod
    def output_manhole_csvformat(cls):
        data_array = np.empty((0,18))
        
        for manhole in cls.manhole_instances.values():
            data_array = np.append(data_array, np.array([manhole.output_info()]), axis=0)
        
        np.set_printoptions(suppress=True)
        # print(data_array)
        # np.savetxt(Path.cwd() / 'inputfiles_edit/MANHOLE.csv', data_array, delimiter=',')
        
         # 途中閲覧用のエクセルファイルを作る。
        # try:
        #     wb = openpyxl.load_workbook(cls.TEMP_XLSX)
        # except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        
        header_lis = ['人孔番号', '人孔断面積', '人孔上の地盤高', '人孔位置のメッシュ番号', '出口フラッグ','排水区番号','集水面積(m2)','不浸透面積','等価粗度','斜面勾配','雨量データ番号','貯留関数β','貯留関数α','ベースカット流量','レベルカット流量','矩形放流量','人孔x座標','人孔y座標']
        for i, header in enumerate(header_lis):
            ws.cell(row=1, column=i+1, value=header)
            
        write_list_2d(ws, data_array, start_row=2, start_col=1)
        
        wb.save(cls.TEMP_XLSX)
        
    
        
class Mesh:
    mesh_instances:dict["Mesh"] = {}
    MESH_WIDTH = 4.988 * 2 # メッシュの横方向長さ。単位:m 今回は縦横合わせて4つを結合して1つのメッシュとする。
    MESH_HEIGHT = 6.184 * 2 # メッシュの縦方向長さ。単位:m    
    third_mesh_infodict = {} # 第三次メッシュの座標などの情報。mesh_protocolsにて代入？  
    start_coord = [0, 0] # メッシュの開始地点の座標。北西端を想定
    map_diagram = [[54372787, 54372788],
                   [54372777, 54372778]]
    

    # NILIMメッシュ
    num_of_vertical:int # 縦の個数
    num_of_horizontal:int # 横の個数
    
    origin_coord:list[float, float] = [0.0, 0.0]

    DEM_CSV = Path.cwd() / r'mesh_dem/dem.csv'
    BRATIO_CSV = Path.cwd() / r'building_ratio/bratio.csv'
    TEMP_XLSX = Path.cwd() / r'inputfiles_edit/MESH.xlsx'
    
    def __init__(self, row, col, dem_data) -> None:
        # self.border_coord_upside:float
        # self.border_coord_downside:float
        # self.border_coord_rightside:float
        # self.border_coord_leftside:float
        self.row_num = row
        self.col_num = col
        self.position = (int(self.col_num), int(self.row_num))
        
        self.dem:float = float(dem_data)
        
        self.manhole_to_supply_water:set = set()
        
        self.building_ratio:float # 建物占有率
        
        self.serial_num:int # nilim用のメッシュ通し番号。
    
    # @classmethod
    # def get_mesh_lengthes(cls):
    
    @classmethod
    def trimming_dem_mesh(cls):
        wb:openpyxl.Workbook = mesh_protocols.load_dem_file()
        dem_lis = []
        for line in cls.map_diagram:
            line_lis = []
            for map_id in line:
                ws = wb[str(map_id)]
                row_lis = []
                for row in ws.iter_rows():
                    temp_lis = []
                    for cell in row:
                        temp_lis.append(cell.value)
                    row_lis.append(temp_lis)
                line_lis.append(row_lis)
            dem_lis.append(line_lis)
        dem_lis = np.array(dem_lis)
        # print(dem_lis) #標高のnp配列が完成。4面分
        
        dem_lis = np.where(dem_lis==-9999, 595, dem_lis)
        
        dem_lis0 = dem_lis[0][0][126:, 9:]
        dem_lis1 = dem_lis[0][1][126:, :(121+1)]
        dem_lis2 = dem_lis[1][0][:(139+1), 9:]
        dem_lis3 = dem_lis[1][1][:(139+1), :(121+1)]
        temp_0 = np.hstack([dem_lis0, dem_lis1])
        temp_1 = np.hstack([dem_lis2, dem_lis3])
        usage_original_area = np.vstack([temp_0, temp_1])
        np.savetxt(Path.cwd()/'mesh_dem/dem_usage_original_area.csv', usage_original_area, delimiter=',')
        # print(usage_original_area) # 対象範囲の標高の2次元配列ができた。
        # print("-----")
        
        # 横2こずつで平均
        cal_evenlis = usage_original_area[:, ::2] 
        cal_oddlis = usage_original_area[:, 1::2]
        
        temp_area = (cal_evenlis + cal_oddlis) / 2
        
        # 縦2個ずつで平均
        cal_evenlis2 = temp_area[::2]
        cal_oddlis2 = temp_area[1::2]
        
        usage_area = (cal_evenlis2 + cal_oddlis2) / 2
        
        
        np.savetxt(Path.cwd()/'mesh_dem/dem_usage_area.csv', usage_area, delimiter=',')
        
        # print(usage_area)
        with open(cls.DEM_CSV,'wt') as f:
            writer = csv.writer(f)
            
            writer.writerows(usage_area)
        
        # test用
        
        # test_array = np.arange(1,33751).reshape(150, 225)
        # test_array = np.array([[test_array,test_array],[test_array, test_array]])
        # # print(test_array)
        # test_array0 = test_array[0][0][126:, 9:]
        # test_array1 = test_array[0][1][126:, :(25+1)]
        # test_array2 = test_array[1][0][:(54+1), 9:]
        # test_array3 = test_array[1][1][:(54+1), :(25+1)]
        # test_0 = np.hstack([test_array0, test_array1])
        
        # print(test_0)
    @classmethod
    def create_mesh_instances(cls): # 保存したcsvから、メッシュインスタンスを生成して、辞書に代入。
        usage_area = np.loadtxt(cls.DEM_CSV, delimiter=',')
        for i, row in enumerate(usage_area):
            # print(row.size)
            for j, data in enumerate(row):
                mesh = Mesh(i, j, data)

                cls.mesh_instances[(j, i)] = mesh
                # print((j,i))
                # if (i, j) == (92, 26):
                #     print((i,j))
        cls.num_of_horizontal = usage_area.shape[1] # メッシュの横に並ぶ個数
        cls.num_of_vertical = usage_area.shape[0]# メッシュの縦に並ぶ個数
        
        # 個数節約のため、一部のメッシュを削除する。
        for j2 in range(46, 82):
            for i2 in range(60, 169): 
                del cls.mesh_instances[(i2, j2)]
        
    @classmethod
    def trimming_b_ratio_csv(cls):# 建物占有率のcsvをメッシュと同じ形にする。
        map_id_lis = [87, 88, 77, 78]
        for map_id in map_id_lis:
            csv_filepath = Path.cwd() / f'building_ratio/b_ratio_{map_id}.csv'
            df = pd.read_csv(csv_filepath, header=0)
            df = df["b_ratio"]
            bratio_array = df.values.reshape(150, 225)
            # print(bratio_array)
            save_filepath = Path.cwd() / f'building_ratio/save_{map_id}.csv'
            np.savetxt(save_filepath, bratio_array, delimiter=',')
            
        array0 = np.loadtxt(Path.cwd() / f'building_ratio/save_87.csv', delimiter=',')
        array1 = np.loadtxt(Path.cwd() / f'building_ratio/save_88.csv', delimiter=',')
        array2 = np.loadtxt(Path.cwd() / f'building_ratio/save_77.csv', delimiter=',')
        array3 = np.loadtxt(Path.cwd() / f'building_ratio/save_78.csv', delimiter=',')
        
        array0_cut = array0[126:, 9:]
        array1_cut = array1[126:, :(121+1)]
        array2_cut = array2[:(139+1), 9:]
        array3_cut = array3[:(139+1), :(121+1)]
        
        temp_0 = np.hstack([array0_cut, array1_cut])
        temp_1 = np.hstack([array2_cut, array3_cut])
        
        usage_original_area = np.vstack([temp_0, temp_1])
        np.savetxt(Path.cwd()/f'building_ratio/usage_original_area_bratio.csv', usage_original_area, delimiter=',')
        
        cal_evenlis = usage_original_area[:, ::2]
        cal_oddlis = usage_original_area[:, 1::2]
        
        temp_area = (cal_evenlis + cal_oddlis) / 2
        
        
        cal_evenlis2 = temp_area[::2]
        cal_oddlis2 = temp_area[1::2]
        
        usage_area = (cal_evenlis2 + cal_oddlis2) / 2
        
        np.savetxt(Path.cwd()/'building_ratio/bratio_usage_area.csv', usage_area, delimiter=',')
        
        with open(cls.BRATIO_CSV,'wt') as f:
            writer = csv.writer(f)
            
            writer.writerows(usage_area)
            

    @classmethod
    def registrate_b_ratio(cls): # 整形した建物占有率のデータを、各メッシュに登録する。
        data_array = np.loadtxt(cls.BRATIO_CSV, delimiter=',')
        # print(data_array[1][3])
        for i, line in enumerate(data_array):
            for j, data in enumerate(line):
                try:
                    mesh:'Mesh' = cls.mesh_instances[(j, i)]
                    mesh.building_ratio = data * 100 # %表示
                except KeyError:
                    pass # 切り取った分のメッシュは無視する。
    
    @classmethod
    def registrate_serial(cls): # 削除されたメッシュを無視して番号を振り分ける。
        serial_num = 1
        for j in range(0, 82):
            for i in range(0, 169):
                try:
                    mesh = cls.mesh_instances[(i, j)]
                    mesh.serial_num = serial_num
                    # print(serial_num)
                    serial_num += 1
                except KeyError:
                    pass
    @classmethod            
    def classify_Manhole(cls, manhole:"Manhole"): # そのマンホールの位置するメッシュのposition（i,j）を返す。
        i = (manhole.coordinates_lis[0] - cls.origin_coord[0]) // cls.MESH_WIDTH
        j = - (manhole.coordinates_lis[1] - cls.origin_coord[1]) // cls.MESH_HEIGHT
        return (int(i), int(j))
    
    @classmethod
    def position_to_id(cls, position:tuple):  #廃止 # position(i,j)から、通し番号に変換。
        # 廃止
        # if (position[0] >= 0) and (position[0] <= (cls.num_of_horizontal-1)) and (position[1] >=0) and (position[1] <= (cls.num_of_vertical - 1)):
        #     return position[1] * cls.num_of_horizontal + position[0] + 1
        # else:
        #     return 0
        
        #(0 <= position[0] <= (cls.num_of_horizontal-1)) and (0 <= position[1] <= (cls.num_of_vertical - 1)) という書き方の方がいいかもしれない。
        pass
    
    @classmethod
    def position_to_serial(cls, position:tuple):
        try:
            return cls.mesh_instances[position].serial_num
        except KeyError:
            return 0
    
    def output_info(self):
        # (position_to_idは廃止)
        # id = Mesh.position_to_id(self.position)
        # north_mesh = Mesh.position_to_id((self.position[0], self.position[1]-1))
        # east_mesh = Mesh.position_to_id((self.position[0]+1, self.position[1]))
        # south_mesh = Mesh.position_to_id((self.position[0], self.position[1]+1))
        # west_mesh = Mesh.position_to_id((self.position[0]-1, self.position[1]))
        
        serial_num = Mesh.position_to_serial(self.position)
        north_mesh = Mesh.position_to_serial((self.position[0], self.position[1]-1))
        east_mesh = Mesh.position_to_serial((self.position[0]+1, self.position[1]))
        south_mesh = Mesh.position_to_serial((self.position[0], self.position[1]+1))
        west_mesh = Mesh.position_to_serial((self.position[0]-1, self.position[1]))
        
        if self.manhole_to_supply_water == 0:
            manhole_to_supply_water = 1
        elif self.manhole_to_supply_water == set():
            manhole_to_supply_water = 0 # これは、集水させないという意味での0。本プログラムの最下流のマンホールの0番とは異なるので注意。
        else:
            manhole_to_supply_water = list(self.manhole_to_supply_water)[0]
        # メッシュデータ
        return np.array([serial_num, 0, north_mesh, east_mesh, south_mesh, west_mesh, self.dem, 0.047, self.building_ratio, 100.0, 1, manhole_to_supply_water, list(Kanro.k_numset)[0], self.position[0], self.position[1]])
    
    @classmethod
    def output_mesh_csvformat(cls):
        data_array = np.empty((0, 15))
        
        
        for mesh in cls.mesh_instances.values():
            # mesh.output_info()
            data_array = np.append(data_array, np.array([mesh.output_info()]), axis=0)
            
        np.set_printoptions(suppress=True)
        # print(data_array)
        # np.savetxt(Path.cwd() / 'inputfiles_edit/Mesh.csv', data_array, delimiter=',')
        
         # 途中閲覧用のエクセルファイルを作る。
        # try:
        #     wb = openpyxl.load_workbook(cls.TEMP_XLSX)
        # except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        
        ws['A1'] = 'メッシュ個数'
        ws['B1'] = 'メッシュ横幅'
        ws['C1'] = 'メッシュ縦幅'
        ws['A2'] = len(cls.mesh_instances)
        ws['B2'] = cls.MESH_WIDTH
        ws['C2'] = cls.MESH_HEIGHT
        
        header_lis = ['メッシュ個数番号','メッシュタイプ','接続メッシュ番号(北側)','(東側)','(南側)','(西側)','平均地盤高(標高m)','建物以外の底面積粗度係数','建物占有率(%)','不浸透面積(%)','雨量観測所番号','集水先人孔番号','排水区番号','二次元I番号','二次元J番号']
        for i, header in enumerate(header_lis):
            ws.cell(row=3, column=i+1, value=header)
            
        write_list_2d(ws, data_array, start_row=4, start_col=1)
        
        wb.save(cls.TEMP_XLSX)
        
        

        

def is_empty2(cell):# 空セルかどうかの判定基準。引用元　-> (https://gammasoft.jp/support/openpyxl-iter-rows/#to-empty-row)の「空行を読み取る」
    return cell.value is None or not str(cell.value).strip()

def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])




if __name__ == '__main__':
    # result_all = Kanro.txt_to_dataframe(KANRO_FILEPATH)
    # Kanro407 = Kanro(1, 407)
    # Kanro407.get_its_result(result_all)
    # Kanro407.display_info()
    # Kanro407.display_graph_hdw() # グラフを表示する。


    Kanro.database_to_list() # staticmethod。はじめに実行する。
    # print(Kanro.database_lis)
    
    # シミュレーションで扱う対象の管路リスト
    registrate_lis = [376, 377, 378, 379, 380, 381, 382, 383, 384, 385, 386, 387, 388, 389, 390, 391, 392, 406, 407,\
                      120, 148, 173, 174, 171, 176, 177, 178, 184, 185, 193, 192, 190, 188, 198, 195, 194, \
                      201, 200, 199, 210, 208, 211, 213, 215, 206, 207, 205, 203, 202, 216, 217, 218, 219, 223, 224,\
                      235, 234, 233, 236, 237, 238, 405, 404, 403, 328, 325, 322, 321, 320, 319, 220, 221, 222,\
                      345, 343, 341,\
                      402, 401, 400, 305, 304, 303, 302, 301, 300, 299, 298, 297, 286, 284, 317, 318, 316, 314, 313, 315,\
                      170, 169, 168, 166, 167, 164, 163, 165, 162, 161, 160, 141, 140, 139, 138, 137, 136, 135, 147,146, 145, 144, 143, 142,\
                      119, 117, 116, 115, 114, 113, 112, 111, 110, 109, 108, 107, 106, 105, 104, 103, 102, 101,\
                      134, 133, 132, 131, 130, 129, 128, 127, 126, 125, 124, 123, 122, 121, 340, 338, 337, 336, 335, 334, 333, 332, 331, 330, 329, 214,\
                      399, 398, 397, 396, 395, 296, 295, 294, 293, 292, 291, 290, 289, 288, 287, 281, 280, 342, 344, 326, 327, 229, 228, 227, 226, 225,\
                      230, 231, 232, 183, 182, 181, 180, 179, 159, 158, 157, 156, 154, 153, 152, 151, 150, 149,\
                      285, 284, 283, 282,\
                      277, 276, 275, 274, 273, 272, 271, 270, 269, 268, 267, 266, 265, 264, 263, 262, 261, 260, 259, 258, 257, 256, 255, 254, 253, 252, 251,\
                      250, 249, 248, 247, 246, 245, 244, 243, 242, 241, 240, 239]
    

    
    # 絞り込み
    Kanro.narrow_down_registration(registrate_lis)
    
    
    # 相互参照
    Kanro.cross_reference()
    
    # 管路インスタンスの生成。排水区ごと。
    for k_num in Kanro.k_numset:
        Kanro.create_kanro_instances(k_num)
        
        
    # print(Kanro.kanro_instances)

    # 最も下流の管路番号を指定する。
    # lowest_pipe_num = 407

    # 管路同士を繋ぎ、マンホールインスタンスを生成する。
    Kanro.start_make_connections()
    
    
    # 管路側に隣接人孔番号が渡っているか確認。
    # for id, pipe in Kanro.kanro_instances.items():
    #     print(id, pipe.neighbor_man_numset)
        
    # 隣接人孔を上流側と下流側で区別。
    # for kanro in Kanro.kanro_instances.values():
    #     kanro.discriminate_manholes()
    #     print(f"管路番号{kanro.pipe_num}| 上流-> {kanro.upstream_mannum} 下流-> {kanro.downstream_mannum}")
    Kanro.start_discrimination()
        
    Kanro.get_coordinates()
    
    Kanro.start_give_coordinates()
    
    Kanro.to_networkx()
    # print(Kanro.kanro_instances[387])
    
   
    
    # 以下、メッシュの作成
    # xmlファイルから、標高データのエクセルを作成する。1度だけ実施すれば良い。
    # mesh_protocols.start_extracting(mesh_protocols.XML_FILEPATH_LIST)
    
    # エクセルから必要な情報（標高データとメッシュの番号タプル）を抽出する。
    Mesh.trimming_dem_mesh()
    
    
    # # CSVに保存した情報から、メッシュのインスタンスを生成。
    Mesh.create_mesh_instances()
    Mesh.registrate_serial()
    
    # # メッシュに建物占有率の情報を与える。
    Mesh.trimming_b_ratio_csv()
    # NILIM用に、メッシュの通し番号を付与する。
    Mesh.registrate_b_ratio()
    
    # # メッシュからマンホールに情報を与える。
    Manhole.start_which_Mesh_and_elevation()
    
    # # マンホールからメッシュに、集水先人孔番号の情報を与える。
    Manhole.start_catch_water_from()
   
        
        
    
    
    
    # # # データ閲覧用ファイルを作成する。
    
    Kanro.output_kanro_csvformat()
    Manhole.output_manhole_csvformat()
    Mesh.output_mesh_csvformat()

    # 調査用
    # manhole199 = Manhole.manhole_instances[210]
    # mesh199:'Mesh' = Mesh.mesh_instances[manhole199.mesh_position]
    # print(mesh199.dem)
    # print(Kanro.kanro_instances[210].ie_up)
    
    
    
    
    
    
    
    
    
    
    """
    main_pipes = [376, 377, 378, 379, 380, 381, 382, 383, 384, 385, 386, 387, 388, 389, 390, 391, 392, 406, 407]
    registrate1 = [120]
    registrate2 = [148]
    registrate3 = [173, 174, 171]
    registrate4 = [176, 177]
    registrate5 = [178]
    registrate6 = [182, 184, 185]
    registrate7 = [193, 192, 190, 188]
    registrate8 = [198, 195, 194]
    registrate9 = [201, 200, 199, 210, 208]
    registrate10 = [211, 213, 215]
    registrate11 = [206, 207]
    registrate12 = [205, 203, 202]
    registrate13 = [216]
    registrate14 = [217, 218]
    registrate15 = [219]
    registrate16 = [223]
    registrate17 = [224]
    registrate18 = [235, 234, 233, 229]
    registrate19 = [236, 237]
    registrate20 = [238]
    registrate21 = [405, 404, 403]
    registrate22 = [328, 325]
    registrate23 = [220, 221, 222]
    registrate24 = [322, 321, 320, 319]
    """
