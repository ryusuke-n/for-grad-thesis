import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib
from matplotlib import dates as mdates
from matplotlib.dates import DateFormatter

from datetime import datetime
from datetime import timedelta

import numpy as np
import openpyxl

from pathlib import Path

import re



def scraping_precipitation(year, month, day):
    url = f'https://www.data.jma.go.jp/obd/stats/etrn/view/10min_s1.php?prec_no=48&block_no=47618&year={year}&month={month}&day={day}&view='
    table = pd.read_html(url)
    # print(len(table))
    df:pd.DataFrame = table[0]

    dates = pd.date_range(f'{year}/{month}/{day} 00:10:00', periods=144, freq='10min') # 日付のインデックスを生成。

    df.index = dates
    df.replace('--', '0', inplace=True)
    df.replace('×', '0', inplace=True)
    df.replace('#', '0', inplace=True)
    df = df.astype({('降水量 (mm)','降水量 (mm)'): float}) # Multiindexの指定はタプルで。
    df = df['降水量 (mm)']['降水量 (mm)'] # 降水量のデータだけ抽出
    df = df * 6 # 6倍して降雨強度(mm/h)に変換する。
    
    # csvに保存
    RAIN_CSV = Path.cwd() / f'precipitations/precipitation_{year}{month:02}{day:02}.csv'
    df.to_csv(RAIN_CSV, header=False, index=True) # 日ごと
    MONTHLY_CSV = Path.cwd() / f'precipitations/monthly/precipitation_{year}{month:02}.csv'
    df.to_csv(MONTHLY_CSV, header=True, index=True, mode='a', encoding='utf-8') # 月間

    
    ###########
    start_datetime = datetime(year, month, day, 0, 10, 0)
    end_datetime = start_datetime + timedelta(days=1)
    x = np.arange(start_datetime, end_datetime, np.timedelta64(10,'m'), dtype='datetime64')
  
    x = x.astype(datetime)
    y = np.array(df.values, dtype=np.float64)
    
    plt.bar(x,y, width=0.01, color='#00AEEF', edgecolor='#00AEEF')
    plt.xticks(rotation=30)
    # plt.xlim(left=np.datetime64('2015-06-26T00:10:00.000000000'), right=np.datetime64('2015-06-27T00:00:00.000000000'))
    plt.ylim(0.0, 15.0)
    
    plt.title(f"降雨データ - {year}{month:02}{day:02}", fontsize=16)
    # plt.xlabel("時刻", fontsize=16)
    plt.ylabel("降雨強度\n(mm/h)", labelpad = 20, fontsize=11, rotation = "horizontal")
    # plt.xticks(fontsize=16)
    plt.yticks(fontsize=12)
    
    plt.subplots_adjust(left=0.15, bottom=0.15)
    
    rain_filepath = Path.cwd().parent / f'rainfalls/charts/Graph_rainfall_{year}{month:02}{day:02}'
    plt.savefig(rain_filepath, dpi=300)
    
    # plt.show()

    
    return df


    
def slice_data(df, start, end):
    # return df[str(start):str(end)]
    pass
    

def to_NILIMformat(df:pd.DataFrame):
    global year
    global month
    global day

    values = df.values
    # values = np.reshape(values, )
    values.resize(18, 8)
    try:
        wb = openpyxl.load_workbook(Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day:02}.xlsx')
    except:
        wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    write_list_2d(ws, values, 3, 1)
    valuesize = values.size
    ws['A1'] = 1
    
    wb.save(Path.cwd() / f'precipitations/rainfall_{year}{month:02}{day:02}.xlsx')
    


def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

def scraping_dailymax_precipitations(year, month): # 日ごとの最大1時間降水量を月間単位で保存する。
    url = f'https://www.data.jma.go.jp/obd/stats/etrn/view/daily_s1.php?prec_no=48&block_no=47618&year={year}&month={month:02}&day=12&view='
    table = pd.read_html(url)
    df = table[0]
    df_dailymax = df['降水量(mm)']['降水量(mm)']['最大']['1時間']

    df_day = df['日']['日']['日']['日']

    df_dailymax = pd.concat([df_day, df_dailymax], axis=1)
    df_dailymax.rename(columns={'1時間':'最大1時間雨量'}, inplace=True)
    
    df_dailymax = df_dailymax.set_index('日')
    
    csv_path = Path.cwd() / f'precipitations/monthly/dailymax/dailymax_{year}{month:02}.csv'
    df_dailymax.to_csv(csv_path, header=True, index=True, encoding='shift_jis')
    
if __name__ == '__main__':
    # year = 2015
    # month = 6
    # day = 30
    # df = scraping_precipitation(year, month, day)
    # print(df)
    # df = slice_data(df, '')
    # to_NILIMformat(df)
    # make_precipitation_graph(df, year, month, day)
    
    
    year = 2019
    # month = 10
    # day_lis = [x for x in range(1, 32)] # 31日までの月はrange(1, 32), 30日までの月はrange(1, 31)となることに注意。
    # for day in day_lis:
    #     df = scraping_precipitation(year, month, day)
    #     to_NILIMformat(df)
    
    for month in range(2, 13):
        scraping_dailymax_precipitations(year, month)
