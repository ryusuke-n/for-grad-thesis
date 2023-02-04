from pathlib import Path, PurePath
import re
from decimal import Decimal

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl

# メッシュの範囲の設定
AREA_SETTINGS = (50, 20)

# App内設定用
WINDOW_SIZE = '400x300'
OUTPUT_FILE = 'output_file.xlsx'# フォーマット成果用。縦一列に並べる時のファイル
SAVEDATA_FILE = 'savedata_file.xlsx'# データ永続用
LOADDATA_FILE = 'loaddata_file.xlsx'# 読み込み用

# 出来上がったcsv or xlsxファイルをtxtにする時のファイル
MESH_TXT = 'kansen_mesh.txt'
MESH_XLSX = 'kansen_mesh_note.xlsx'

class MeshApp(ttk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=380, height=280,
                         borderwidth=1, relief='groove')
        self.root = root
        self.root.title("メッシュ操作ツール")
        
        self.output_filename = Path.cwd() / OUTPUT_FILE
        self.savedata_filename = Path.cwd() / SAVEDATA_FILE
        self.loaddata_filename = Path.cwd() / LOADDATA_FILE
        
        self.CreateMesh(AREA_SETTINGS[0], AREA_SETTINGS[1]) # テスト用
        self.SaveData() # テスト用
        

    def CreateMesh(self, row_nums, col_nums):
        # 保存ボタン
        save_btn = tk.Button(self.root, width=2, height=1)
        save_btn['text'] = "保存"
        save_btn['command'] = self.SaveData
        save_btn.grid(row=1, column=1, columnspan=2)
        
        # 読み込みボタン
        load_btn = tk.Button(self.root, width=2, height=1)
        load_btn['text'] = '読み込み'
        load_btn['command'] = self.LoadData
        load_btn.grid(row=1, column=3, columnspan=2)
        
        # 出力ボタン
        output_btn = tk.Button(self.root, width=2, height=1)
        output_btn['text'] = '出力'
        output_btn['command'] = self.OutputData
        output_btn.grid(row=1, column=5, columnspan=2)
        
        # 閉じるボタン
        quit_btn = tk.Button(self.root, width=2, height=1)
        quit_btn['text'] = "閉じる"
        quit_btn['command'] = self.root.destroy
        quit_btn.grid(row=1, column=7, columnspan=2)
        # メッシュマトリックスの作成
        self.matrix_lis = []
        for i in range(col_nums):
            row_lis = []
            for j in range(row_nums):
                
                mesh = tk.Entry(self.root, width=2)
                mesh.grid(row=(i + 4), column=(j + 1))
                
                row_lis.append(mesh)
            self.matrix_lis.append(row_lis)
         
    
    def SaveData(self): # 記入内容をエクセルシートに保存する工程。
        wb = openpyxl.load_workbook(self.savedata_filename)
        ws = wb.worksheets[0]
        
        for i in range(len(self.matrix_lis)):
            for j in range(len(self.matrix_lis[i])):
                text = self.matrix_lis[i][j].get()
                position = str(num2alpha(j+1)) + str(i + 1)
                ws[position].value = text # positionでエクセル上の位置を指定して、textの中身を書き込む。
        wb.save(self.savedata_filename)
    
    def LoadData(self):
        wb = openpyxl.load_workbook(self.loaddata_filename)
        ws = wb.worksheets[0]
        
        for i in range(len(self.matrix_lis)):
            for j in range(len(self.matrix_lis[i])):
                position = str(num2alpha(j+1)) + str(i + 1)
                # print(position, type(position))
                if (value := ws[position].value) is None:# セルが空白の場合は0を代入する。
                    # テキストボックスに0を記入する
                    self.matrix_lis[i][j].insert(0, "0")
                    
                    # print(value)
                else:
                    # テキストボックスにvalueを記入する
                    self.matrix_lis[i][j].insert(0, str(value))
                    # print(value)
                    
    def OutputData(self):# NILIM2.0の求めるフォーマットの下書き(note)を作る
        wb = openpyxl.load_workbook(self.output_filename)
        ws = wb.worksheets[0]
        
        count = 0
        for i in range(len(self.matrix_lis)):
            for j in range(len(self.matrix_lis[i])):
                position = "L" + str(count + 4)
                if (value := self.matrix_lis[i][j].get()) is None:
                    ws[position].value = 0
                    count += 1
                else:
                    ws[position].value = int(value)
                    count += 1
                    
        wb.save(self.output_filename)
    
    @staticmethod
    def CreateMeshFormat(input_filename, output_filename):
        input_filepath = Path.cwd().parent / input_filename
        output_filepath = Path.cwd().parent / output_filename
        
        f_txt = open(output_filepath, 'w', newline="")
        
        wb = openpyxl.load_workbook(input_filepath)
        ws = wb.worksheets[0]

        for row in ws.iter_rows(min_row=4):
            if all(is_empty(c) for c in row):
                break
            else:
                temp_lis = []
                for cell in row:
                    value = cell.value
                    print(value)
                    col_num = cell.column
                    # print(type(value))
                    # print(col_num, type(col_num))
                    # if col_num  == 1 or 2:
                    #     temp_lis.append(f'{value:>5}')
                    if col_num == 3 or 4 or 5 or 6:
                        temp_lis.append(f'{value:>10}')
                    elif col_num == 7:
                        temp_lis.append(f'{value:>10.2f}')
                    elif col_num == 8:
                        temp_lis.append(f'{value:>10.4f}')
                    elif col_num == 9 or 10:
                        temp_lis.append(f'{value:>10.2f}')
                    elif col_num == 11 or 12 or 13 or 14:
                        temp_lis.append(f'{value:>10}')
                    elif col_num == 15:
                        temp_lis.append(f'{value:>10}')
                    
                    print("--turn end--")
                # for cell in row:
                #     n_string = cell.value
                # print(temp_lis)
                n_string = "".join(temp_lis)
                # print(n_string)
                f_txt.write(str(n_string) + "\n")
                
        f_txt.close()
                
                

                
# 数値→アルファベットを27以降でも行う関数。参考→https://tanuhack.com/num2alpha-alpha2num/
# この関数を使わなくても、openpyxlに.cordinateメソッドや.columnメソッドを使えば一発かも？
def num2alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num%26)
    
# アルファベット→数値
def alpha2num(alpha):
    num=0
    for index, item in enumerate(list(alpha)):
        num += pow(26,len(alpha)-index-1)*(ord(item)-ord('A')+1)
    return num

def position2index(alpha_num:str):
    alpha = re.sub(r"[^a-zA-Z]", "", alpha_num) # エクセルのセル位置からアルファベットを抽出。正規表現を利用
    num = re.sub(r"[^0-9]", "", alpha_num)# エクセルのセル位置からアルファベットを抽出。正規表現を利用。
    row_index = num - 1 # エクセルは1はじまり、Pythonのindexは0はじまり。
    col_index = alpha2num(alpha) - 1
    return row_index, col_index

def is_empty(cell):# 空セルかどうかの判定基準。引用元　-> (https://gammasoft.jp/support/openpyxl-iter-rows/#to-empty-row)の「空行を読み取る」
    return cell.value is None or not str(cell.value).strip()
    


if __name__ == '__main__':
    root = tk.Tk()
    app = MeshApp(root=root)
    root.geometry(WINDOW_SIZE)
    app.mainloop()
    
    # MeshApp.CreateMeshFormat(MESH_XLSX, MESH_TXT)
    
    # print(position2index("AAL302"))